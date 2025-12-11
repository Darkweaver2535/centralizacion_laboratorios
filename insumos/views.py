from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Count
from django.core.paginator import Paginator
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
import json
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .models import Insumo
from .forms import InsumoForm
from core.models import UnidadAcademica, Laboratorio, Carrera, Asignatura, UnidadTematica, GuiaLaboratorio, Practica


def lista_insumos(request):
    """Vista principal de insumos con filtros y paginación"""
    
    # Obtener parámetros de filtro
    unidad_academica_id = request.GET.get('unidad_academica')
    laboratorio_id = request.GET.get('laboratorio')
    categoria = request.GET.get('categoria')
    estado = request.GET.get('estado')
    carrera_id = request.GET.get('carrera')
    nombre_elemento = request.GET.get('nombre_elemento')
    
    # Consulta base
    insumos = Insumo.objects.select_related(
        'unidad_academica', 'laboratorio', 'carrera', 'asignatura', 'unidad_tematica'
    ).all()
    
    # Aplicar filtros
    if unidad_academica_id:
        insumos = insumos.filter(unidad_academica_id=unidad_academica_id)
    
    if laboratorio_id:
        insumos = insumos.filter(laboratorio_id=laboratorio_id)
    
    if categoria:
        insumos = insumos.filter(categoria=categoria)
    
    if estado:
        insumos = insumos.filter(estado=estado)
    
    if carrera_id:
        insumos = insumos.filter(carrera_id=carrera_id)
    
    if nombre_elemento:
        insumos = insumos.filter(
            Q(nombre_elemento__icontains=nombre_elemento) |
            Q(descripcion_caracteristicas__icontains=nombre_elemento)
        )
    
    # Ordenar por fecha de creación (más recientes primero)
    insumos = insumos.order_by('-id')
    
    # Estadísticas
    total_insumos = insumos.count()
    disponibles = insumos.filter(estado__in=['nuevo', 'bueno']).count()
    bajo_stock = insumos.filter(cantidad__lt=5).count()
    agotados = insumos.filter(cantidad=0).count()
    
    stats = {
        'total': total_insumos,
        'disponibles': disponibles,
        'bajo_stock': bajo_stock,
        'agotados': agotados,
    }
    
    # Datos para filtros
    unidades_academicas = UnidadAcademica.objects.all()
    laboratorios = Laboratorio.objects.all()
    carreras = Carrera.objects.all()
    
    context = {
        'insumos': insumos,
        'stats': stats,
        'unidades_academicas': unidades_academicas,
        'laboratorios': laboratorios,
        'carreras': carreras,
        'title': 'Gestión de Insumos',
        'subtitle': 'Control de inventario y administración de insumos de laboratorio'
    }
    
    return render(request, 'insumos/lista.html', context)


@login_required
def importar_insumos_view(request):
    """Vista para importar insumos desde Excel/CSV"""
    return render(request, 'insumos/importar.html', {
        'tipo': 'insumos'
    })

def nuevo_insumo(request):
    """Vista para crear un nuevo insumo"""
    
    if request.method == 'POST':
        form = InsumoForm(request.POST)
        if form.is_valid():
            insumo = form.save(commit=False)
            # Asignar el usuario creador si está autenticado
            if request.user.is_authenticated:
                insumo.usuario_creador = request.user
            insumo.save()
            messages.success(request, 'Insumo agregado correctamente.')
            return redirect('insumos:lista')
        else:
            # Mostrar errores específicos para debugging
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
            messages.error(request, 'Error al crear el insumo. Revise los errores arriba.')
    else:
        form = InsumoForm()
    
    context = {
        'form': form,
        'title': 'Nuevo Insumo',
        'subtitle': 'Registrar un nuevo insumo en el inventario'
    }
    
    return render(request, 'insumos/nuevo.html', context)


def detalle_insumo(request, insumo_id):
    """Vista de detalle de un insumo específico"""
    
    insumo = get_object_or_404(Insumo, id=insumo_id)
    
    context = {
        'insumo': insumo,
        'title': f'Detalle: {insumo.nombre_elemento}',
        'subtitle': f'Código: {insumo.codigo_inventario}'
    }
    
    return render(request, 'insumos/detalle.html', context)


def editar_insumo(request, insumo_id):
    """Vista para editar un insumo existente"""
    
    insumo = get_object_or_404(Insumo, id=insumo_id)
    
    if request.method == 'POST':
        form = InsumoForm(request.POST, instance=insumo)
        if form.is_valid():
            insumo = form.save()
            messages.success(request, 'Insumo actualizado correctamente.')
            return redirect('insumos:detalle', insumo_id=insumo.id)
        else:
            messages.error(request, 'Error al actualizar el insumo. Verifique los datos ingresados.')
    else:
        form = InsumoForm(instance=insumo)
    
    context = {
        'form': form,
        'insumo': insumo,
        'title': f'Editar: {insumo.nombre_elemento}',
        'subtitle': f'Modificar información del insumo'
    }
    
    return render(request, 'insumos/editar.html', context)


@require_http_methods(["DELETE"])
def eliminar_insumo(request, insumo_id):
    """Vista para eliminar un insumo"""
    
    try:
        insumo = get_object_or_404(Insumo, id=insumo_id)
        nombre = insumo.nombre_elemento
        insumo.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Insumo eliminado correctamente.'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al eliminar el insumo: {str(e)}'
        })


def exportar_excel(request):
    """Exportar insumos a Excel con las 19 columnas oficiales"""
    
    # Aplicar los mismos filtros que en la vista principal
    unidad_academica_id = request.GET.get('unidad_academica')
    laboratorio_id = request.GET.get('laboratorio')
    categoria = request.GET.get('categoria')
    estado = request.GET.get('estado')
    carrera_id = request.GET.get('carrera')
    nombre_elemento = request.GET.get('nombre_elemento')
    
    insumos = Insumo.objects.select_related(
        'unidad_academica', 'laboratorio', 'carrera', 'asignatura', 'unidad_tematica'
    ).all()
    
    # Aplicar filtros
    if unidad_academica_id:
        insumos = insumos.filter(unidad_academica_id=unidad_academica_id)
    if laboratorio_id:
        insumos = insumos.filter(laboratorio_id=laboratorio_id)
    if categoria:
        insumos = insumos.filter(categoria=categoria)
    if estado:
        insumos = insumos.filter(estado=estado)
    if carrera_id:
        insumos = insumos.filter(carrera_id=carrera_id)
    if nombre_elemento:
        insumos = insumos.filter(
            Q(nombre_elemento__icontains=nombre_elemento) |
            Q(descripcion_caracteristicas__icontains=nombre_elemento)
        )
    
    # Crear libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario de Insumos"
    
    # Encabezados de las 19 columnas oficiales
    headers = [
        'UNIDAD ACADÉMICA',
        'LABORATORIO',
        'CATEGORÍA',
        'NOMBRE DEL ELEMENTO',
        'DESCRIPCIÓN/CARACTERÍSTICAS',
        'MARCA/MODELO',
        'CÓDIGO DE INVENTARIO',
        'ESTADO',
        'UBICACIÓN FÍSICA',
        'CANTIDAD',
        'UNIDAD DE MEDIDA',
        'FECHA DE INGRESO/COMPRA',
        'USO PRINCIPAL',
        'CARRERA',
        'ASIGNATURA',
        'UNIDAD TEMÁTICA',
        'CONDICIONES DE ALMACENAMIENTO',
        'OBSERVACIONES',
        'LINK DE LA FOTOGRAFÍA'
    ]
    
    # Escribir encabezados
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Escribir datos
    for row, insumo in enumerate(insumos, 2):
        data = [
            insumo.unidad_academica.nombre if insumo.unidad_academica else '',
            insumo.laboratorio.nombre if insumo.laboratorio else '',
            insumo.get_categoria_display(),
            insumo.nombre_elemento or '',
            insumo.descripcion_caracteristicas or '',
            insumo.marca_modelo or '',
            insumo.codigo_inventario or '',
            insumo.get_estado_display(),
            insumo.ubicacion_fisica or '',
            insumo.cantidad or 0,
            insumo.unidad_medida or '',
            insumo.fecha_ingreso_compra.strftime('%d/%m/%Y') if insumo.fecha_ingreso_compra else '',
            insumo.get_uso_principal_display() if insumo.uso_principal else '',
            insumo.carrera.nombre if insumo.carrera else '',
            insumo.asignatura.nombre if insumo.asignatura else '',
            insumo.unidad_tematica.nombre if insumo.unidad_tematica else '',
            insumo.get_condiciones_almacenamiento_display() if insumo.condiciones_almacenamiento else '',
            insumo.observaciones or '',
            insumo.link_fotografia or ''
        ]
        
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
    
    # Ajustar ancho de columnas
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 15
    
    # Preparar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"inventario_insumos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


# APIs para dropdowns dinámicos

def api_carreras(request):
    """API para obtener carreras filtradas por unidad académica"""
    
    unidad_academica = request.GET.get('unidad_academica')
    
    if unidad_academica:
        # Mapear los valores del formulario a los nombres en la base de datos
        mapeo_unidades = {
            'la_paz': 'UALP',
            'santa_cruz': 'UASC', 
            'cochabamba': 'UACB',
            'riberalta': 'UCRB',
            'tropico': 'UATP'
        }
        
        # Obtener la unidad académica por ID o por nombre mapeado
        unidad = None
        
        # Si es un número, buscar por ID
        if unidad_academica.isdigit():
            try:
                unidad = UnidadAcademica.objects.get(id=int(unidad_academica))
            except UnidadAcademica.DoesNotExist:
                return JsonResponse({'error': 'Unidad académica no encontrada'}, status=404)
        else:
            # Si es texto, mapear a nombre oficial
            nombre_unidad = mapeo_unidades.get(unidad_academica)
            if nombre_unidad:
                try:
                    unidad = UnidadAcademica.objects.get(nombre=nombre_unidad)
                except UnidadAcademica.DoesNotExist:
                    return JsonResponse({'error': f'Unidad académica {nombre_unidad} no encontrada'}, status=404)
        
        if unidad:
            carreras = Carrera.objects.filter(unidad_academica=unidad)
        else:
            carreras = Carrera.objects.all()
    else:
        carreras = Carrera.objects.all()
    
    data = {
        'carreras': [
            {'id': carrera.id, 'nombre': carrera.get_nombre_display()}
            for carrera in carreras
        ]
    }
    
    return JsonResponse(data)


def api_asignaturas(request):
    """API para obtener asignaturas filtradas por carrera y semestre"""
    
    carrera_id = request.GET.get('carrera')
    semestre = request.GET.get('semestre')
    
    if carrera_id:
        asignaturas = Asignatura.objects.filter(carrera_id=carrera_id)
        
        # Si se especifica semestre, filtrar también por semestre
        if semestre:
            try:
                semestre_int = int(semestre)
                asignaturas = asignaturas.filter(semestre=semestre_int)
            except ValueError:
                pass  # Si el semestre no es un número válido, ignorar el filtro
    else:
        asignaturas = Asignatura.objects.all()
    
    data = {
        'asignaturas': [
            {
                'id': asignatura.id, 
                'nombre': asignatura.nombre,
                'semestre': asignatura.semestre
            }
            for asignatura in asignaturas.order_by('nombre')
        ]
    }
    
    return JsonResponse(data)


def api_unidades_tematicas(request):
    """API para obtener unidades temáticas filtradas por asignatura"""
    
    asignatura_id = request.GET.get('asignatura')
    
    if asignatura_id:
        unidades = UnidadTematica.objects.filter(asignatura_id=asignatura_id)
    else:
        unidades = UnidadTematica.objects.all()
    
    data = {
        'unidades_tematicas': [
            {'id': unidad.id, 'nombre': unidad.nombre}
            for unidad in unidades.order_by('numero', 'nombre')
        ]
    }
    
    return JsonResponse(data)


def api_guias_laboratorio(request):
    """API para obtener guías de laboratorio filtradas por unidad temática"""
    
    unidad_tematica_id = request.GET.get('unidad_tematica')
    
    if unidad_tematica_id:
        guias = GuiaLaboratorio.objects.filter(unidad_tematica_id=unidad_tematica_id)
    else:
        guias = GuiaLaboratorio.objects.all()
    
    data = {
        'guias_laboratorio': [
            {'id': guia.id, 'nombre': guia.nombre}
            for guia in guias.order_by('numero', 'nombre')
        ]
    }
    
    return JsonResponse(data)


def api_practicas(request):
    """API para obtener prácticas filtradas por guía de laboratorio"""
    
    guia_laboratorio_id = request.GET.get('guia_laboratorio')
    
    if guia_laboratorio_id:
        practicas = Practica.objects.filter(guia_laboratorio_id=guia_laboratorio_id)
    else:
        practicas = Practica.objects.all()
    
    data = {
        'practicas': [
            {'id': practica.id, 'nombre': practica.nombre}
            for practica in practicas.order_by('numero', 'nombre')
        ]
    }
    
    return JsonResponse(data)
