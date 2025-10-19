from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.db.models import Count, Q
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator
from equipos.models import Equipo
from insumos.models import Insumo
from core.models import UnidadAcademica, Carrera, Laboratorio, GuiaLaboratorio, Asignatura, UnidadTematica, UnidadDidactica, ContenidoAnalitico, PracticaLaboratorio
from guias.models import GuiaGenerada
import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.utils import timezone
import datetime

def aplicar_filtros_equipos(equipos, params):
    """Aplicar filtros jerárquicos a equipos"""
    if params.get('unidad_academica'):
        equipos = equipos.filter(unidad_academica_id=params.get('unidad_academica'))
    if params.get('carrera'):
        equipos = equipos.filter(carrera_id=params.get('carrera'))
    if params.get('semestre'):
        equipos = equipos.filter(semestre=params.get('semestre'))
    if params.get('asignatura'):
        equipos = equipos.filter(asignatura_id=params.get('asignatura'))
    if params.get('unidad_didactica'):
        equipos = equipos.filter(unidad_didactica_id=params.get('unidad_didactica'))
    if params.get('contenido_analitico'):
        equipos = equipos.filter(contenido_analitico_id=params.get('contenido_analitico'))
    if params.get('laboratorio'):
        equipos = equipos.filter(laboratorio_id=params.get('laboratorio'))
    if params.get('estado'):
        equipos = equipos.filter(estado=params.get('estado'))
    if params.get('responsable'):
        equipos = equipos.filter(responsable_excel__icontains=params.get('responsable'))
    if params.get('busqueda'):
        equipos = equipos.filter(
            Q(equipo_existente__icontains=params.get('busqueda')) |
            Q(marca__icontains=params.get('busqueda')) |
            Q(modelo__icontains=params.get('busqueda')) |
            Q(responsable_excel__icontains=params.get('busqueda'))
        )
    return equipos

def aplicar_filtros_insumos(insumos, params):
    """Aplicar filtros jerárquicos a insumos"""
    if params.get('unidad_academica'):
        insumos = insumos.filter(unidad_academica=params.get('unidad_academica'))
    if params.get('carrera'):
        insumos = insumos.filter(carrera=params.get('carrera'))
    # Nota: Los insumos no tienen campo semestre directo
    # if params.get('semestre'):
    #     insumos = insumos.filter(semestre=params.get('semestre'))
    if params.get('asignatura'):
        insumos = insumos.filter(asignatura=params.get('asignatura'))
    if params.get('laboratorio'):
        insumos = insumos.filter(laboratorio=params.get('laboratorio'))
    if params.get('busqueda'):
        insumos = insumos.filter(
            Q(nombre_elemento__icontains=params.get('busqueda')) |
            Q(descripcion_caracteristicas__icontains=params.get('busqueda')) |
            Q(codigo_inventario__icontains=params.get('busqueda'))
        )
    return insumos

def aplicar_filtros_guias(practicas, params):
    """Aplicar filtros jerárquicos a prácticas de laboratorio (PracticaLaboratorio)"""
    if params.get('unidad_academica'):
        practicas = practicas.filter(contenido_analitico__unidad_didactica__asignatura__carrera__icontains=params.get('unidad_academica'))
    if params.get('carrera'):
        practicas = practicas.filter(contenido_analitico__unidad_didactica__asignatura__carrera__icontains=params.get('carrera'))
    if params.get('semestre'):
        practicas = practicas.filter(contenido_analitico__unidad_didactica__asignatura__semestre=params.get('semestre'))
    if params.get('asignatura'):
        practicas = practicas.filter(contenido_analitico__unidad_didactica__asignatura__nombre__icontains=params.get('asignatura'))
    if params.get('unidad_didactica'):
        practicas = practicas.filter(contenido_analitico__unidad_didactica__nombre__icontains=params.get('unidad_didactica'))
    if params.get('busqueda'):
        practicas = practicas.filter(
            Q(nombre__icontains=params.get('busqueda')) |
            Q(contenido_analitico__descripcion__icontains=params.get('busqueda')) |
            Q(contenido_analitico__unidad_didactica__nombre__icontains=params.get('busqueda')) |
            Q(contenido_analitico__unidad_didactica__asignatura__nombre__icontains=params.get('busqueda'))
        )
    return practicas

@login_required
def visualizacion_view(request):
    """Vista R2 unificada para visualización de Equipos, Insumos y Guías con django-filter integrado"""
    
    # Importar filtros de django-filter
    from .filters import EquipoFilter, InsumoFilter, GuiaFilter, PracticaLaboratorioFilter
    from guias.models import GuiaGenerada
    
    # Obtener categoría seleccionada y validarla
    categoria = request.GET.get('categoria', 'equipos')
    
    # Validar categoría y usar por defecto si es inválida
    categorias_validas = ['equipos', 'insumos', 'guias']
    if categoria not in categorias_validas or categoria == 'undefined':
        categoria = 'equipos'  # Categoría por defecto
    
    # Inicializar variables
    items = None
    filterset = None
    stats = {}
    
    # Lógica según categoría seleccionada - USAR DJANGO-FILTER
    if categoria == 'equipos':
        # Usar EquipoFilter en lugar de lógica manual
        filterset = EquipoFilter(request.GET, queryset=Equipo.objects.select_related(
            'unidad_academica', 'carrera', 'asignatura',
            'criterio_desempeno', 'unidad_didactica', 'contenido_analitico',
            'guia_laboratorio', 'practica', 'laboratorio'
        ))
        items = filterset.qs  # Queryset filtrado por django-filter
        
        # Estadísticas para equipos
        stats = {
            'total_items': items.count(),
            'total_laboratorios': Laboratorio.objects.count(),
            'items_buenos': items.filter(estado='bueno').count(),
            'items_regulares': items.filter(estado='regular').count(),
            'items_malos': items.filter(estado='malo').count(),
            'categoria_nombre': 'Equipos'
        }
        
    elif categoria == 'insumos':
        # Usar InsumoFilter en lugar de lógica manual
        filterset = InsumoFilter(request.GET, queryset=Insumo.objects.select_related(
            'unidad_academica', 'carrera', 'asignatura', 'laboratorio'
        ))
        items = filterset.qs  # Queryset filtrado por django-filter
        
        # Estadísticas para insumos
        stats = {
            'total_items': items.count(),
            'total_laboratorios': Laboratorio.objects.count(),
            'items_buenos': items.filter(estado='disponible').count() if hasattr(Insumo, 'estado') else 0,
            'items_regulares': items.filter(estado='agotado').count() if hasattr(Insumo, 'estado') else 0,
            'items_malos': items.filter(estado='vencido').count() if hasattr(Insumo, 'estado') else 0,
            'categoria_nombre': 'Insumos'
        }
        
    elif categoria == 'guias':
        # Usar PracticaLaboratorioFilter para filtrado automático
        filterset = PracticaLaboratorioFilter(request.GET, queryset=PracticaLaboratorio.objects.select_related(
            'contenido_analitico__unidad_didactica__asignatura'
        ).prefetch_related(
            'contenido_analitico__competencias',
            'contenido_analitico__objetivos_practica'
        ).order_by('contenido_analitico__unidad_didactica__asignatura__carrera', 
                   'contenido_analitico__unidad_didactica__asignatura__semestre', 
                   'orden'))
        items = filterset.qs  # Queryset filtrado por django-filter
        
        # Estadísticas para guías (prácticas)
        stats = {
            'total_items': items.count(),
            'total_laboratorios': Laboratorio.objects.count(),
            'items_buenos': items.count(),  # Todas las prácticas son "buenas"
            'items_regulares': 0,
            'items_malos': 0,
            'categoria_nombre': 'Guías de Laboratorio (Prácticas)'
        }
    
    # Verificación de seguridad para items
    if items is None:
        # Si por alguna razón items es None, usar queryset vacío por defecto
        from django.db.models import QuerySet
        items = Equipo.objects.none()  # Queryset vacío pero válido
        stats = {
            'total_items': 0,
            'total_laboratorios': Laboratorio.objects.count(),
            'items_buenos': 0,
            'items_regulares': 0,
            'items_malos': 0,
            'categoria_nombre': 'Sin categoría'
        }
    
    # Paginación unificada
    paginator = Paginator(items, 50)  # 50 items por página
    page_number = request.GET.get('page')
    items_page = paginator.get_page(page_number)
    
    # Datos para filtros jerárquicos (comunes a todas las categorías)
    unidades = UnidadAcademica.objects.all()
    carreras = Carrera.objects.all()
    asignaturas = Asignatura.objects.all()
    unidades_tematicas = UnidadTematica.objects.all()
    laboratorios = Laboratorio.objects.all()
    
    # Obtener responsables únicos según categoría
    responsables = []
    if categoria == 'equipos':
        responsables = Equipo.objects.exclude(responsable_excel='').values_list('responsable_excel', flat=True).distinct().order_by('responsable_excel')
    
    # Choices para dropdowns
    semestres_choices = [(i, f'{i}°') for i in range(1, 11)]
    
    # Estados según categoría
    estados_choices = []
    if categoria == 'equipos':
        estados_choices = [
            ('bueno', 'Bueno'),
            ('regular', 'Regular'),
            ('malo', 'Malo'),
        ]
    elif categoria == 'insumos':
        estados_choices = [
            ('disponible', 'Disponible'),
            ('agotado', 'Agotado'),
            ('vencido', 'Vencido'),
        ]
    
    # Obtener correlaciones (relaciones entre equipos, insumos y guías)
    correlaciones = obtener_correlaciones(categoria, items_page, request.GET)
    
    # Context específico por categoría
    context = {
        # Datos comunes - SIEMPRE mostrar totales reales para la interfaz principal
        'categoria': categoria,
        'stats': {
            'total_equipos': Equipo.objects.count(),
            'total_insumos': Insumo.objects.count(),
            'total_guias': PracticaLaboratorio.objects.count(),  # Usar PracticaLaboratorio
            'equipos_buenos': Equipo.objects.filter(estado='bueno').count(),
            'equipos_regulares': Equipo.objects.filter(estado='regular').count(),
            'equipos_malos': Equipo.objects.filter(estado='malo').count(),
        },
        
        # Filtros jerárquicos
        'unidades_academicas': unidades,
        'carreras': carreras,
        'asignaturas': asignaturas,
        'unidades_didacticas': unidades_tematicas,
        'contenidos_analiticos': [],  # TODO: Agregar si existe este modelo
        'semestres': range(1, 11),
        
        # NUEVA: Información de filtros django-filter
        'filterset': filterset,  # Para acceder a los filtros en el template
        'filtered_count': items.count(),  # Número de elementos filtrados
        'total_count': filterset.queryset.model.objects.count() if filterset else 0,  # Total sin filtros
        
        # Valores seleccionados para mantener en formulario
        'unidad_seleccionada': request.GET.get('unidad_academica', ''),
        'carrera_seleccionada': request.GET.get('carrera', ''),
        'semestre_seleccionado': request.GET.get('semestre', ''),
        'asignatura_seleccionada': request.GET.get('asignatura', ''),
        'unidad_didactica_seleccionada': request.GET.get('unidad_didactica', ''),
        'contenido_seleccionado': request.GET.get('contenido', ''),
        
        # Correlaciones
        'correlaciones': correlaciones,
    }
    
    # Agregar datos específicos según categoría
    if categoria == 'equipos':
        context['equipos'] = items_page
    elif categoria == 'insumos':
        context['insumos'] = items_page
    elif categoria == 'guias':
        # Para guías, pasar directamente los items sin paginación por ahora
        context['guias'] = items  # Usar items directamente
        context['items'] = items  # También agregar como items para el template
    
    return render(request, 'visualizacion_r2.html', context)

def obtener_correlaciones(categoria, items, filtros):
    """Obtener correlaciones entre equipos, insumos y guías según la categoría seleccionada"""
    correlaciones = {
        'equipos_relacionados': [],
        'insumos_relacionados': [],
        'guias_relacionadas': []
    }
    
    if categoria == 'equipos' and items:
        # Para equipos: mostrar insumos y guías relacionadas
        for equipo in items[:5]:  # Limitar para rendimiento
            if hasattr(equipo, 'guia_laboratorio') and equipo.guia_laboratorio:
                correlaciones['guias_relacionadas'].append({
                    'nombre': equipo.guia_laboratorio.nombre,
                    'asignatura': equipo.asignatura.get_nombre_display() if equipo.asignatura else 'N/A'
                })
    
    elif categoria == 'insumos' and items:
        # Para insumos: mostrar equipos y guías que los usan
        for insumo in items[:5]:  # Limitar para rendimiento
            if hasattr(insumo, 'asignatura') and insumo.asignatura:
                correlaciones['equipos_relacionados'].append({
                    'nombre': f'Equipos de {insumo.asignatura.get_nombre_display()}',
                    'carrera': insumo.carrera.get_nombre_display() if insumo.carrera else 'N/A'
                })
    
    elif categoria == 'guias' and items:
        # Para guías: mostrar equipos e insumos relacionados
        for guia in items[:5]:  # Limitar para rendimiento
            if hasattr(guia, 'unidad_tematica') and guia.unidad_tematica:
                correlaciones['equipos_relacionados'].append({
                    'nombre': f'Equipos para {guia.nombre}',
                    'asignatura': guia.unidad_tematica.asignatura.get_nombre_display() if guia.unidad_tematica.asignatura else 'N/A'
                })
    
    return correlaciones

@login_required
def filtrar_datos(request):
    """Vista AJAX para filtrar datos"""
    return JsonResponse({
        'status': 'success',
        'message': 'Filtros aplicados correctamente'
    })

@login_required
def obtener_opciones_filtro(request):
    """Vista AJAX para obtener opciones de filtro"""
    return JsonResponse({
        'status': 'success',
        'opciones': []
    })

@login_required
def equipos_ajax(request):
    """Vista AJAX para obtener equipos"""
    return JsonResponse({
        'status': 'success',
        'equipos': []
    })

# ===============================================
# ENDPOINTS AJAX PARA FILTROS DINÁMICOS R2
# ===============================================

@login_required
def ajax_carreras_por_unidad(request):
    """Endpoint AJAX para obtener carreras por unidad académica"""
    unidad_id = request.GET.get('unidad_id')
    
    if not unidad_id:
        return JsonResponse({'carreras': []})
    
    try:
        carreras = Carrera.objects.filter(unidad_academica_id=unidad_id).values(
            'id', 'nombre'
        ).order_by('nombre')
        
        carreras_list = [{
            'id': carrera['id'],
            'nombre': carrera['nombre'],
            'display': dict(Carrera.CARRERAS).get(carrera['nombre'], carrera['nombre'])
        } for carrera in carreras]
        
        return JsonResponse({
            'success': True,
            'carreras': carreras_list
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
def ajax_asignaturas_por_carrera(request):
    """Endpoint AJAX para obtener asignaturas por carrera"""
    carrera_id = request.GET.get('carrera_id')
    
    if not carrera_id:
        return JsonResponse({'asignaturas': []})
    
    try:
        asignaturas = Asignatura.objects.filter(carrera_id=carrera_id).values(
            'id', 'nombre', 'semestre'
        ).order_by('semestre', 'nombre')
        
        asignaturas_list = [{
            'id': asignatura['id'],
            'nombre': asignatura['nombre'],
            'semestre': asignatura['semestre'],
            'display': f"{dict(Asignatura.ASIGNATURAS_CHOICES).get(asignatura['nombre'], asignatura['nombre'])} ({asignatura['semestre']}° Sem)"
        } for asignatura in asignaturas]
        
        return JsonResponse({
            'success': True,
            'asignaturas': asignaturas_list
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
def ajax_semestres_por_carrera(request):
    """Endpoint AJAX para obtener semestres disponibles por carrera"""
    carrera_id = request.GET.get('carrera_id')
    
    if not carrera_id:
        return JsonResponse({'semestres': []})
    
    try:
        semestres = Asignatura.objects.filter(
            carrera_id=carrera_id
        ).values_list('semestre', flat=True).distinct().order_by('semestre')
        
        semestres_list = [{
            'numero': sem,
            'display': f'{sem}° Semestre'
        } for sem in semestres]
        
        return JsonResponse({
            'success': True,
            'semestres': semestres_list
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
def ajax_unidades_didacticas_por_asignatura(request):
    """Endpoint AJAX para obtener unidades didácticas por asignatura"""
    asignatura_id = request.GET.get('asignatura_id')
    
    if not asignatura_id:
        return JsonResponse({'unidades_didacticas': []})
    
    try:
        # Obtener unidades temáticas relacionadas con la asignatura
        unidades = UnidadTematica.objects.filter(
            asignatura_id=asignatura_id
        ).values('id', 'nombre', 'numero').order_by('numero')
        
        unidades_list = [{
            'id': unidad['id'],
            'nombre': unidad['nombre'],
            'numero': unidad['numero'],
            'display': f"Unidad {unidad['numero']}: {unidad['nombre']}"
        } for unidad in unidades]
        
        return JsonResponse({
            'success': True,
            'unidades_didacticas': unidades_list
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
def ajax_contenidos_por_unidad_didactica(request):
    """Endpoint AJAX para obtener contenidos analíticos por unidad didáctica"""
    unidad_didactica_id = request.GET.get('unidad_didactica_id')
    
    if not unidad_didactica_id:
        return JsonResponse({'contenidos': []})
    
    try:
        # Por ahora simulamos contenidos - esto se puede expandir con un modelo específico
        contenidos_simulados = [
            {'id': 1, 'nombre': 'Introducción y conceptos básicos'},
            {'id': 2, 'nombre': 'Desarrollo teórico'},
            {'id': 3, 'nombre': 'Aplicaciones prácticas'},
            {'id': 4, 'nombre': 'Evaluación y análisis'},
        ]
        
        return JsonResponse({
            'success': True,
            'contenidos': contenidos_simulados
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
def ajax_estadisticas_filtradas(request):
    """Endpoint AJAX para obtener estadísticas según filtros aplicados"""
    categoria = request.GET.get('categoria', 'equipos')
    
    # Obtener filtros
    filtros = {
        'unidad_academica': request.GET.get('unidad_academica'),
        'carrera': request.GET.get('carrera'),
        'semestre': request.GET.get('semestre'),
        'asignatura': request.GET.get('asignatura'),
    }
    
    try:
        if categoria == 'equipos':
            queryset = Equipo.objects.all()
            queryset = aplicar_filtros_equipos(queryset, filtros)
            
            stats = {
                'total': queryset.count(),
                'buenos': queryset.filter(estado='bueno').count(),
                'regulares': queryset.filter(estado='regular').count(),
                'malos': queryset.filter(estado='malo').count(),
            }
            
        elif categoria == 'insumos':
            queryset = Insumo.objects.all()
            queryset = aplicar_filtros_insumos(queryset, filtros)
            
            stats = {
                'total': queryset.count(),
                'disponibles': queryset.count(),  # Simplificado por ahora
                'poco_stock': 0,
                'agotados': 0,
            }
            
        elif categoria == 'guias':
            from guias.models import GuiaGenerada
            queryset = GuiaGenerada.objects.all()
            
            # Aplicar filtros básicos a guías
            if filtros['carrera']:
                queryset = queryset.filter(carrera_id=filtros['carrera'])
            if filtros['asignatura']:
                queryset = queryset.filter(asignatura_id=filtros['asignatura'])
            
            stats = {
                'total': queryset.count(),
                'aprobadas': queryset.filter(estado='aprobada').count(),
                'borradores': queryset.filter(estado='borrador').count(),
                'revision': queryset.filter(estado='revision').count(),
            }
        
        return JsonResponse({
            'success': True,
            'stats': stats
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
def exportar_excel(request):
    """Vista para exportar datos a Excel"""
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="equipos.xlsx"'
    
    # Crear workbook simple
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Equipos"
    
    # Headers
    headers = ['ID', 'Equipo', 'Marca', 'Modelo', 'Estado']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Datos
    equipos = Equipo.objects.all()[:100]  # Limitar a 100 para evitar problemas
    for row, equipo in enumerate(equipos, 2):
        ws.cell(row=row, column=1, value=equipo.id)
        ws.cell(row=row, column=2, value=equipo.equipo_existente)
        ws.cell(row=row, column=3, value=equipo.marca or '')
        ws.cell(row=row, column=4, value=equipo.modelo or '')
        ws.cell(row=row, column=5, value=equipo.get_estado_display())
    
    wb.save(response)
    return response


# ===============================================
# SISTEMA DE CORRELACIONES CRÍTICO
# ===============================================

def ajax_correlaciones_equipo(request):
    """Devuelve las correlaciones de un equipo específico con guías e insumos"""
    try:
        equipo_id = request.GET.get('equipo_id')
        if not equipo_id:
            return JsonResponse({'success': False, 'error': 'equipo_id requerido'})
            
        equipo = get_object_or_404(Equipo, id=equipo_id)
        
        # Obtener guías que requieren este equipo
        guias_relacionadas = GuiaGenerada.objects.filter(equipos_requeridos=equipo)
        guias_data = []
        for guia in guias_relacionadas:
            guias_data.append({
                'id': guia.id,
                'titulo': guia.titulo,
                'carrera': guia.carrera.get_nombre_display(),
                'asignatura': guia.asignatura.get_nombre_display(),
                'estado': guia.estado,
                'semestre': guia.semestre,
            })
        
        # Obtener insumos relacionados a través de las guías
        insumos_relacionados = set()
        for guia in guias_relacionadas:
            for insumo in guia.insumos_requeridos.all():
                insumos_relacionados.add(insumo)
        
        insumos_data = []
        for insumo in insumos_relacionados:
            insumos_data.append({
                'id': insumo.id,
                'nombre': insumo.nombre_elemento,
                'categoria': insumo.get_categoria_display(),
                'estado': insumo.estado,
                'cantidad': insumo.cantidad,
            })
            
        return JsonResponse({
            'success': True,
            'equipo': {
                'id': equipo.id,
                'nombre': equipo.equipo_existente,
                'estado': equipo.estado,
                'categoria': equipo.get_categoria_display(),
            },
            'correlaciones': {
                'guias_relacionadas': guias_data,
                'insumos_relacionados': insumos_data,
                'total_guias': len(guias_data),
                'total_insumos': len(insumos_data),
            }
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


def ajax_correlaciones_guia(request):
    """Devuelve las correlaciones de una guía específica con equipos e insumos"""
    try:
        guia_id = request.GET.get('guia_id')
        if not guia_id:
            return JsonResponse({'success': False, 'error': 'guia_id requerido'})
            
        guia = get_object_or_404(GuiaGenerada, id=guia_id)
        
        # Obtener equipos requeridos por esta guía
        equipos_data = []
        for equipo in guia.equipos_requeridos.all():
            equipos_data.append({
                'id': equipo.id,
                'nombre': equipo.equipo_existente,
                'estado': equipo.estado,
                'categoria': equipo.get_categoria_display(),
                'laboratorio': equipo.laboratorio.get_nombre_display() if equipo.laboratorio else 'Sin asignar',
            })
        
        # Obtener insumos requeridos por esta guía
        insumos_data = []
        for insumo in guia.insumos_requeridos.all():
            insumos_data.append({
                'id': insumo.id,
                'nombre': insumo.nombre_elemento,
                'categoria': insumo.get_categoria_display(),
                'estado': insumo.estado,
                'cantidad': insumo.cantidad,
                'unidad': insumo.unidad_medida,
            })
            
        return JsonResponse({
            'success': True,
            'guia': {
                'id': guia.id,
                'titulo': guia.titulo,
                'estado': guia.estado,
                'carrera': guia.carrera.get_nombre_display(),
                'asignatura': guia.asignatura.get_nombre_display(),
            },
            'correlaciones': {
                'equipos_requeridos': equipos_data,
                'insumos_requeridos': insumos_data,
                'total_equipos': len(equipos_data),
                'total_insumos': len(insumos_data),
            }
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


def ajax_resumen_correlaciones(request):
    """Devuelve un resumen general de todas las correlaciones del sistema"""
    try:
        # Estadísticas generales de correlaciones
        total_guias = GuiaGenerada.objects.count()
        guias_con_equipos = GuiaGenerada.objects.filter(equipos_requeridos__isnull=False).distinct().count()
        guias_con_insumos = GuiaGenerada.objects.filter(insumos_requeridos__isnull=False).distinct().count()
        
        total_equipos = Equipo.objects.count()
        equipos_en_guias = Equipo.objects.filter(guiagenerada__isnull=False).distinct().count()
        
        # Correlaciones más frecuentes (equipos más utilizados)
        from django.db.models import Count
        equipos_populares = Equipo.objects.annotate(
            uso_count=Count('guiagenerada')
        ).filter(uso_count__gt=0).order_by('-uso_count')[:5]
        
        equipos_populares_data = []
        for equipo in equipos_populares:
            equipos_populares_data.append({
                'id': equipo.id,
                'nombre': equipo.equipo_existente,
                'usos': equipo.uso_count,
                'estado': equipo.estado,
            })
        
        return JsonResponse({
            'success': True,
            'resumen': {
                'total_guias': total_guias,
                'guias_con_equipos': guias_con_equipos,
                'guias_con_insumos': guias_con_insumos,
                'total_equipos': total_equipos,
                'equipos_en_guias': equipos_en_guias,
                'porcentaje_equipos_utilizados': round((equipos_en_guias / total_equipos * 100), 2) if total_equipos > 0 else 0,
                'equipos_mas_utilizados': equipos_populares_data,
            }
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


def ajax_correlaciones_insumo(request):
    """Devuelve las correlaciones de un insumo específico"""
    try:
        insumo_id = request.GET.get('insumo_id')
        if not insumo_id:
            return JsonResponse({'success': False, 'error': 'ID de insumo requerido'})
        
        from insumos.models import Insumo
        
        # Obtener el insumo específico
        insumo = Insumo.objects.get(id=insumo_id)
        
        # Obtener guías que requieren este insumo
        guias_relacionadas = GuiaGenerada.objects.filter(
            insumos_requeridos=insumo
        ).select_related(
            'unidad_tematica__asignatura__carrera__unidad_academica'
        ).prefetch_related(
            'equipos_requeridos'
        )
        
        # Estructurar datos de las guías relacionadas
        guias_data = []
        total_equipos_relacionados = set()
        
        for guia in guias_relacionadas:
            equipos_guia = list(guia.equipos_requeridos.all())
            equipos_data = []
            
            for equipo in equipos_guia:
                equipos_data.append({
                    'id': equipo.id,
                    'nombre': equipo.equipo_existente,
                    'estado': equipo.estado,
                    'laboratorio': equipo.laboratorio.nombre if equipo.laboratorio else 'N/A'
                })
                total_equipos_relacionados.add(equipo.id)
            
            guia_info = {
                'id': guia.id,
                'titulo': guia.titulo,
                'descripcion': guia.descripcion[:100] + '...' if len(guia.descripcion) > 100 else guia.descripcion,
                'asignatura': guia.unidad_tematica.asignatura.nombre,
                'carrera': guia.unidad_tematica.asignatura.carrera.nombre,
                'equipos_relacionados': equipos_data,
                'total_equipos': len(equipos_data)
            }
            guias_data.append(guia_info)
        
        # Información del insumo
        insumo_info = {
            'id': insumo.id,
            'nombre': insumo.nombre_elemento,
            'categoria': insumo.categoria,
            'cantidad': insumo.cantidad,
            'unidad_medida': insumo.unidad_medida,
            'estado': insumo.estado,
            'descripcion': insumo.descripcion_caracteristicas,
            'laboratorio': insumo.laboratorio.nombre if insumo.laboratorio else 'N/A'
        }
        
        return JsonResponse({
            'success': True,
            'insumo': insumo_info,
            'guias_relacionadas': guias_data,
            'total_guias': len(guias_data),
            'total_equipos_relacionados': len(total_equipos_relacionados),
            'resumen': {
                'impacto_guias': len(guias_data),
                'impacto_equipos': len(total_equipos_relacionados),
                'categoria': insumo.categoria,
                'disponibilidad': insumo.estado
            }
        })
        
    except Insumo.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Insumo no encontrado'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def exportar_excel_avanzado(request):
    """Exportar datos a Excel con formato profesional y correlaciones completas"""
    try:
        from .exportacion_utils import exportar_excel_completo
        
        categoria = request.GET.get('categoria', 'equipos')
        
        # Obtener parámetros de filtros
        filtros_params = {
            'unidad_academica': request.GET.get('unidad_academica'),
            'carrera': request.GET.get('carrera'),
            'semestre': request.GET.get('semestre'),
            'asignatura': request.GET.get('asignatura'),
            'laboratorio': request.GET.get('laboratorio'),
            'estado': request.GET.get('estado'),
            'busqueda': request.GET.get('busqueda'),
        }
        
        # Aplicar filtros según categoría
        if categoria == 'equipos':
            datos = Equipo.objects.all()
            datos = aplicar_filtros_equipos(datos, filtros_params)
            datos = datos.select_related('laboratorio', 'carrera', 'asignatura').prefetch_related('guiagenerada_set__insumos_requeridos')
        elif categoria == 'insumos':
            datos = Insumo.objects.all()
            datos = aplicar_filtros_insumos(datos, filtros_params)
            datos = datos.select_related('laboratorio', 'carrera', 'asignatura').prefetch_related('guiagenerada_set__equipos_requeridos')
        elif categoria == 'guias':
            datos = GuiaGenerada.objects.all()
            datos = aplicar_filtros_guias(datos, filtros_params)
            datos = datos.select_related('carrera', 'asignatura').prefetch_related('equipos_requeridos', 'insumos_requeridos')
        else:
            return JsonResponse({'success': False, 'error': 'Categoría no válida'})
        
        # Filtrar parámetros no vacíos para mostrar en resumen
        filtros_aplicados = {k: v for k, v in filtros_params.items() if v}
        
        return exportar_excel_completo(categoria, list(datos), filtros_aplicados)
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error al exportar: {str(e)}'})


@login_required 
def exportar_pdf_guia(request):
    """Exportar guía específica a PDF para docentes"""
    try:
        from .exportacion_utils import exportar_pdf_guia_completa
        
        guia_id = request.GET.get('guia_id')
        if not guia_id:
            return JsonResponse({'success': False, 'error': 'ID de guía requerido'})
        
        response = exportar_pdf_guia_completa(guia_id)
        
        if response is None:
            return JsonResponse({'success': False, 'error': 'Guía no encontrada'})
        
        return response
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error al exportar PDF: {str(e)}'})


@login_required
def exportar_guias_filtradas_pdf(request):
    """Exportar múltiples guías filtradas en un solo PDF"""
    try:
        from .exportacion_utils import exportar_pdf_guia_completa
        import io
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate
        from PyPDF2 import PdfMerger
        
        # Obtener parámetros de filtros
        filtros_params = {
            'unidad_academica': request.GET.get('unidad_academica'),
            'carrera': request.GET.get('carrera'),
            'semestre': request.GET.get('semestre'),
            'asignatura': request.GET.get('asignatura'),
        }
        
        # Obtener guías filtradas
        guias = GuiaGenerada.objects.all()
        guias = aplicar_filtros_guias(guias, filtros_params)
        guias = guias.select_related('carrera', 'asignatura').prefetch_related('equipos_requeridos', 'insumos_requeridos')
        
        if not guias:
            return JsonResponse({'success': False, 'error': 'No se encontraron guías con los filtros aplicados'})
        
        # Si es solo una guía, usar exportación individual
        if guias.count() == 1:
            return exportar_pdf_guia_completa(guias.first().id)
        
        # Para múltiples guías, crear un PDF combinado (versión simplificada)
        from datetime import datetime
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.platypus import Paragraph, Spacer, PageBreak
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        story = []
        styles = getSampleStyleSheet()
        
        # Portada
        story.append(Paragraph("GUÍAS DE LABORATORIO EXPORTADAS", styles['Title']))
        story.append(Paragraph(f"Fecha de exportación: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
        story.append(Paragraph(f"Total de guías: {guias.count()}", styles['Normal']))
        
        # Filtros aplicados
        if any(filtros_params.values()):
            story.append(Spacer(1, 12))
            story.append(Paragraph("Filtros aplicados:", styles['Heading2']))
            for filtro, valor in filtros_params.items():
                if valor:
                    story.append(Paragraph(f"• {filtro}: {valor}", styles['Normal']))
        
        story.append(PageBreak())
        
        # Resumen de guías
        story.append(Paragraph("RESUMEN DE GUÍAS", styles['Heading1']))
        for i, guia in enumerate(guias, 1):
            story.append(Paragraph(f"{i}. {guia.titulo}", styles['Heading2']))
            story.append(Paragraph(f"Carrera: {guia.carrera}", styles['Normal']))
            story.append(Paragraph(f"Asignatura: {guia.asignatura}", styles['Normal']))
            story.append(Paragraph(f"Equipos requeridos: {guia.equipos_requeridos.count()}", styles['Normal']))
            story.append(Paragraph(f"Insumos requeridos: {guia.insumos_requeridos.count()}", styles['Normal']))
            story.append(Spacer(1, 12))
        
        doc.build(story)
        buffer.seek(0)
        
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="guias_laboratorio_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        
        return response
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error al exportar PDF: {str(e)}'})


# ========== NUEVOS ENDPOINTS API PARA VISTA DINÁMICA ==========

@login_required
def api_filtros(request, filtro):
    """API para obtener opciones de filtros jerárquicos"""
    try:
        padre = request.GET.get('padre')
        categoria = request.GET.get('categoria', 'equipos')
        
        opciones = []
        
        if filtro == 'carrera' and padre:
            opciones = list(Carrera.objects.filter(unidad_academica_id=padre).values('id', 'nombre'))
        elif filtro == 'semestre' and padre:
            # Obtener semestres únicos según la categoría
            if categoria == 'equipos':
                semestres = Equipo.objects.filter(carrera_id=padre).values_list('semestre', flat=True).distinct()
            elif categoria == 'insumos':
                semestres = Insumo.objects.filter(carrera_id=padre).values_list('semestre', flat=True).distinct()
            else:  # guias
                semestres = GuiaGenerada.objects.filter(carrera_id=padre).values_list('semestre', flat=True).distinct()
            
            opciones = [{'id': sem, 'nombre': f'{sem}° Semestre'} for sem in sorted(semestres) if sem]
        elif filtro == 'asignatura' and padre:
            # Obtener asignaturas del semestre
            carrera_id = request.GET.get('carrera')
            semestre = request.GET.get('semestre')
            if carrera_id and semestre:
                opciones = list(Asignatura.objects.filter(
                    carrera_id=carrera_id, 
                    semestre=semestre
                ).values('id', 'nombre'))
        elif filtro == 'unidad_didactica' and padre:
            opciones = list(UnidadDidactica.objects.filter(asignatura_id=padre).values('id', 'nombre'))
        elif filtro == 'contenido' and padre:
            opciones = list(ContenidoAnalitico.objects.filter(unidad_didactica_id=padre).values('id', 'nombre'))
        
        return JsonResponse(opciones, safe=False)
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required 
def api_buscar(request):
    """API para buscar elementos según filtros - AHORA USA DJANGO-FILTER"""
    try:
        from .filters import EquipoFilter, InsumoFilter, GuiaFilter
        from guias.models import GuiaGenerada
        
        categoria = request.GET.get('categoria', 'equipos')
        
        if categoria == 'equipos':
            # Usar EquipoFilter para filtrado automático
            filterset = EquipoFilter(request.GET, queryset=Equipo.objects.select_related(
                'unidad_academica', 'carrera', 'asignatura', 
                'unidad_didactica', 'contenido_analitico'
            ))
            queryset = filterset.qs
            
            resultados = []
            for equipo in queryset:
                resultados.append({
                    'id': equipo.id,
                    'nombre': equipo.equipo_existente or 'Sin nombre',
                    'descripcion': f'{equipo.marca} {equipo.modelo}'.strip() or 'Sin descripción',
                    'unidad_academica': equipo.unidad_academica.nombre if equipo.unidad_academica else 'N/A',
                    'carrera': equipo.carrera.nombre if equipo.carrera else 'N/A',
                    'semestre': f'{equipo.semestre}° Semestre' if equipo.semestre else 'N/A',
                    'asignatura': equipo.asignatura.nombre if equipo.asignatura else 'N/A',
                    'estado': equipo.estado or 'Sin estado'
                })
                
        elif categoria == 'insumos':
            # Usar InsumoFilter para filtrado automático
            filterset = InsumoFilter(request.GET, queryset=Insumo.objects.select_related(
                'unidad_academica', 'carrera', 'asignatura', 'laboratorio'
            ))
            queryset = filterset.qs
            
            resultados = []
            for insumo in queryset:
                resultados.append({
                    'id': insumo.id,
                    'nombre': insumo.nombre_elemento or 'Sin nombre',
                    'descripcion': insumo.descripcion_caracteristicas or 'Sin descripción',
                    'unidad_academica': insumo.unidad_academica.nombre if insumo.unidad_academica else 'N/A',
                    'carrera': insumo.carrera.nombre if insumo.carrera else 'N/A',
                    'semestre': 'N/A',  # Los insumos no tienen campo semestre
                    'asignatura': insumo.asignatura.nombre if insumo.asignatura else 'N/A',
                    'categoria': insumo.categoria or 'Sin categoría'
                })
                
        else:  # guias
            # Usar GuiaFilter para filtrado automático
            filterset = GuiaFilter(request.GET, queryset=GuiaGenerada.objects.select_related(
                'carrera__unidad_academica', 'asignatura'
            ))
            queryset = filterset.qs
            
            resultados = []
            for guia in queryset:
                resultados.append({
                    'id': guia.id,
                    'nombre': guia.titulo or 'Sin título',
                    'descripcion': guia.contenido_analitico[:100] + '...' if guia.contenido_analitico else 'Sin descripción',
                    'unidad_academica': guia.carrera.unidad_academica.nombre if guia.carrera and guia.carrera.unidad_academica else 'N/A',
                    'carrera': guia.carrera.nombre if guia.carrera else 'N/A',
                    'semestre': f'{guia.semestre}° Semestre' if guia.semestre else 'N/A',
                    'asignatura': guia.asignatura.nombre if guia.asignatura else 'N/A',
                    'tipo': guia.get_tipo_practica_display() if guia.tipo_practica else 'Guía de práctica'
                })
        
        return JsonResponse({
            'count': len(resultados),
            'results': resultados
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def api_categoria(request, categoria):
    """API para obtener información de una categoría"""
    try:
        stats = {}
        
        if categoria == 'equipos':
            stats['total'] = Equipo.objects.count()
            stats['activos'] = Equipo.objects.filter(estado='Operativo').count()
        elif categoria == 'insumos':
            stats['total'] = Insumo.objects.count()
            stats['categorias'] = Insumo.objects.values('categoria').distinct().count()
        elif categoria == 'guias':
            stats['total'] = GuiaGenerada.objects.count()
            stats['por_asignatura'] = GuiaGenerada.objects.values('asignatura').distinct().count()
        
        return JsonResponse({
            'categoria': categoria,
            'stats': stats
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def debug_api_view(request):
    """Vista temporal para debug de APIs"""
    return render(request, 'debug_api.html')
