import io
import pandas as pd
from django.http import HttpResponse
from django.template.loader import get_template
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
from datetime import datetime

def exportar_excel_completo(categoria, datos_filtrados, filtros_aplicados=None):
    """
    Exportar datos a Excel con formato profesional
    """
    
    # Crear workbook
    wb = Workbook()
    
    # Configurar estilos
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
    data_font = Font(name='Arial', size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    if categoria == 'equipos':
        ws = wb.active
        ws.title = "Equipos de Laboratorio"
        
        # Headers
        headers = [
            'ID', 'Nombre del Equipo', 'Marca', 'Modelo', 'Estado',
            'Laboratorio', 'Responsable', 'Carrera', 'Asignatura',
            'Guías que lo Usan', 'Total Guías', 'Fecha Registro'
        ]
        
        # Escribir headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Escribir datos
        for row_num, equipo in enumerate(datos_filtrados, 2):
            # Obtener guías relacionadas
            guias_relacionadas = equipo.guiagenerada_set.all()
            guias_nombres = "; ".join([g.titulo for g in guias_relacionadas])
            
            row_data = [
                equipo.id,
                equipo.equipo_existente,
                equipo.marca or 'N/A',
                equipo.modelo or 'N/A',
                equipo.estado,
                str(equipo.laboratorio) if equipo.laboratorio else 'N/A',
                equipo.responsable_excel or 'N/A',
                str(equipo.carrera) if equipo.carrera else 'N/A',
                str(equipo.asignatura) if equipo.asignatura else 'N/A',
                guias_nombres or 'Ninguna',
                len(guias_relacionadas),
                equipo.fecha_registro.strftime('%d/%m/%Y') if equipo.fecha_registro else 'N/A'
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.font = data_font
                cell.border = border
        
        # Agregar hoja de correlaciones
        ws_corr = wb.create_sheet(title="Correlaciones Equipos")
        
        # Headers para correlaciones
        corr_headers = ['Equipo', 'Guía', 'Asignatura', 'Carrera', 'Insumos Relacionados']
        for col_num, header in enumerate(corr_headers, 1):
            cell = ws_corr.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Escribir correlaciones
        row_corr = 2
        for equipo in datos_filtrados:
            guias = equipo.guiagenerada_set.all()
            for guia in guias:
                insumos_guia = guia.insumos_requeridos.all()
                insumos_nombres = "; ".join([i.nombre_elemento for i in insumos_guia])
                
                corr_data = [
                    equipo.equipo_existente,
                    guia.titulo,
                    str(guia.asignatura) if guia.asignatura else 'N/A',
                    str(guia.carrera) if guia.carrera else 'N/A',
                    insumos_nombres or 'Ninguno'
                ]
                
                for col_num, value in enumerate(corr_data, 1):
                    cell = ws_corr.cell(row=row_corr, column=col_num, value=value)
                    cell.font = data_font
                    cell.border = border
                
                row_corr += 1
    
    elif categoria == 'insumos':
        ws = wb.active
        ws.title = "Insumos de Laboratorio"
        
        # Headers
        headers = [
            'ID', 'Nombre', 'Categoría', 'Cantidad', 'Unidad',
            'Estado', 'Laboratorio', 'Carrera', 'Asignatura',
            'Guías que lo Requieren', 'Total Guías', 'Equipos Relacionados'
        ]
        
        # Escribir headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Escribir datos
        for row_num, insumo in enumerate(datos_filtrados, 2):
            # Obtener guías relacionadas
            guias_relacionadas = insumo.guiagenerada_set.all()
            guias_nombres = "; ".join([g.titulo for g in guias_relacionadas])
            
            # Obtener equipos relacionados a través de guías
            equipos_relacionados = set()
            for guia in guias_relacionadas:
                for equipo in guia.equipos_requeridos.all():
                    equipos_relacionados.add(equipo.equipo_existente)
            equipos_nombres = "; ".join(equipos_relacionados)
            
            row_data = [
                insumo.id,
                insumo.nombre_elemento,
                insumo.categoria,
                insumo.cantidad,
                insumo.unidad_medida,
                insumo.estado,
                str(insumo.laboratorio) if insumo.laboratorio else 'N/A',
                str(insumo.carrera) if insumo.carrera else 'N/A',
                str(insumo.asignatura) if insumo.asignatura else 'N/A',
                guias_nombres or 'Ninguna',
                len(guias_relacionadas),
                equipos_nombres or 'Ninguno'
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.font = data_font
                cell.border = border
    
    elif categoria == 'guias':
        ws = wb.active
        ws.title = "Guías de Laboratorio"
        
        # Headers principales
        headers = [
            'ID', 'Título', 'Carrera', 'Semestre', 'Asignatura',
            'Tipo Práctica', 'Duración (hrs)', 'Estado',
            'Total Equipos', 'Total Insumos'
        ]
        
        # Escribir headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Escribir datos
        for row_num, guia in enumerate(datos_filtrados, 2):
            row_data = [
                guia.id,
                guia.titulo,
                str(guia.carrera) if guia.carrera else 'N/A',
                guia.semestre,
                str(guia.asignatura) if guia.asignatura else 'N/A',
                guia.get_tipo_practica_display(),
                guia.duracion_horas,
                guia.get_estado_display() if hasattr(guia, 'get_estado_display') else 'N/A',
                guia.equipos_requeridos.count(),
                guia.insumos_requeridos.count()
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.font = data_font
                cell.border = border
        
        # Agregar hoja detallada de cada guía
        for guia in datos_filtrados:
            ws_guia = wb.create_sheet(title=f"Guía {guia.id}")
            
            # Información de la guía
            ws_guia.cell(row=1, column=1, value="INFORMACIÓN DE LA GUÍA").font = Font(size=14, bold=True)
            ws_guia.cell(row=2, column=1, value="Título:").font = Font(bold=True)
            ws_guia.cell(row=2, column=2, value=guia.titulo)
            ws_guia.cell(row=3, column=1, value="Objetivo:").font = Font(bold=True)
            ws_guia.cell(row=3, column=2, value=guia.objetivo_general)
            
            # Equipos requeridos
            row_actual = 6
            ws_guia.cell(row=row_actual, column=1, value="EQUIPOS REQUERIDOS").font = Font(size=12, bold=True)
            row_actual += 1
            
            for equipo in guia.equipos_requeridos.all():
                ws_guia.cell(row=row_actual, column=1, value=f"• {equipo.equipo_existente}")
                ws_guia.cell(row=row_actual, column=2, value=f"Estado: {equipo.estado}")
                ws_guia.cell(row=row_actual, column=3, value=f"Lab: {equipo.laboratorio}")
                row_actual += 1
            
            # Insumos requeridos
            row_actual += 2
            ws_guia.cell(row=row_actual, column=1, value="INSUMOS REQUERIDOS").font = Font(size=12, bold=True)
            row_actual += 1
            
            for insumo in guia.insumos_requeridos.all():
                ws_guia.cell(row=row_actual, column=1, value=f"• {insumo.nombre_elemento}")
                ws_guia.cell(row=row_actual, column=2, value=f"{insumo.cantidad} {insumo.unidad_medida}")
                ws_guia.cell(row=row_actual, column=3, value=f"Estado: {insumo.estado}")
                row_actual += 1
    
    # Agregar hoja de resumen
    ws_resumen = wb.create_sheet(title="Resumen Exportación")
    ws_resumen.cell(row=1, column=1, value="RESUMEN DE EXPORTACIÓN").font = Font(size=14, bold=True)
    ws_resumen.cell(row=2, column=1, value=f"Fecha de exportación: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    ws_resumen.cell(row=3, column=1, value=f"Categoría: {categoria.title()}")
    ws_resumen.cell(row=4, column=1, value=f"Total registros: {len(datos_filtrados)}")
    
    if filtros_aplicados:
        ws_resumen.cell(row=6, column=1, value="FILTROS APLICADOS").font = Font(size=12, bold=True)
        row_filtros = 7
        for filtro, valor in filtros_aplicados.items():
            if valor:
                ws_resumen.cell(row=row_filtros, column=1, value=f"{filtro}: {valor}")
                row_filtros += 1
    
    # Ajustar ancho de columnas
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Preparar respuesta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="laboratorio_{categoria}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    # Guardar en respuesta
    wb.save(response)
    
    return response


def exportar_pdf_guia_completa(guia_id):
    """
    Exportar guía completa a PDF para docentes
    """
    from guias.models import GuiaGenerada
    
    try:
        guia = GuiaGenerada.objects.get(id=guia_id)
    except GuiaGenerada.DoesNotExist:
        return None
    
    # Crear documento PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    # Obtener estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        textColor=colors.darkblue,
        alignment=1  # Centrado
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=12,
        spaceBefore=20,
        spaceAfter=10,
        textColor=colors.darkblue
    )
    
    # Contenido del PDF
    story = []
    
    # Título
    story.append(Paragraph(f"GUÍA DE LABORATORIO", title_style))
    story.append(Paragraph(f"{guia.titulo}", styles['Heading1']))
    story.append(Spacer(1, 12))
    
    # Información general
    story.append(Paragraph("INFORMACIÓN GENERAL", heading_style))
    
    info_data = [
        ['Carrera:', str(guia.carrera)],
        ['Semestre:', guia.get_semestre_display()],
        ['Asignatura:', str(guia.asignatura)],
        ['Tipo de Práctica:', guia.get_tipo_practica_display()],
        ['Duración:', f"{guia.duracion_horas} horas"],
        ['Número de Práctica:', str(guia.numero_practica)]
    ]
    
    info_table = Table(info_data, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(info_table)
    story.append(Spacer(1, 20))
    
    # Competencias
    story.append(Paragraph("COMPETENCIAS", heading_style))
    story.append(Paragraph(guia.competencias, styles['Normal']))
    story.append(Spacer(1, 12))
    
    # Objetivo
    story.append(Paragraph("OBJETIVO GENERAL", heading_style))
    story.append(Paragraph(guia.objetivo_general, styles['Normal']))
    story.append(Spacer(1, 12))
    
    # Objetivos específicos
    if guia.objetivos_especificos:
        story.append(Paragraph("OBJETIVOS ESPECÍFICOS", heading_style))
        story.append(Paragraph(guia.objetivos_especificos, styles['Normal']))
        story.append(Spacer(1, 12))
    
    # Equipos requeridos
    equipos = guia.equipos_requeridos.all()
    if equipos:
        story.append(Paragraph("EQUIPOS REQUERIDOS", heading_style))
        
        equipos_data = [['Equipo', 'Estado', 'Laboratorio', 'Responsable']]
        for equipo in equipos:
            equipos_data.append([
                equipo.equipo_existente,
                equipo.estado.title(),
                str(equipo.laboratorio) if equipo.laboratorio else 'N/A',
                equipo.responsable_excel or 'N/A'
            ])
        
        equipos_table = Table(equipos_data, colWidths=[2.5*inch, 1*inch, 1.5*inch, 1.5*inch])
        equipos_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        
        story.append(equipos_table)
        story.append(Spacer(1, 20))
    
    # Insumos requeridos
    insumos = guia.insumos_requeridos.all()
    if insumos:
        story.append(Paragraph("INSUMOS REQUERIDOS", heading_style))
        
        insumos_data = [['Insumo', 'Cantidad', 'Unidad', 'Categoría', 'Estado']]
        for insumo in insumos:
            insumos_data.append([
                insumo.nombre_elemento,
                str(insumo.cantidad),
                insumo.unidad_medida,
                insumo.categoria.title(),
                insumo.estado.title()
            ])
        
        insumos_table = Table(insumos_data, colWidths=[2*inch, 0.8*inch, 0.8*inch, 1*inch, 0.8*inch])
        insumos_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        
        story.append(insumos_table)
        story.append(Spacer(1, 20))
    
    # Procedimientos
    story.append(Paragraph("PROCEDIMIENTOS", heading_style))
    story.append(Paragraph(guia.procedimientos, styles['Normal']))
    story.append(Spacer(1, 12))
    
    # Referencias bibliográficas
    if guia.referencia_bibliografica:
        story.append(Paragraph("REFERENCIAS BIBLIOGRÁFICAS", heading_style))
        story.append(Paragraph(guia.referencia_bibliografica, styles['Normal']))
    
    # Construir PDF
    doc.build(story)
    
    # Preparar respuesta
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="guia_{guia.id}_{guia.titulo.replace(" ", "_")}.pdf"'
    
    return response