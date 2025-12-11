from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.urls import reverse
from django.core.exceptions import ValidationError
import pandas as pd
import openpyxl
from io import BytesIO
import mimetypes

@login_required
def seleccionar_metodo_carga(request, tipo):
    """
    Vista para seleccionar el método de carga (manual vs Excel/CSV)
    tipo puede ser: 'equipos' o 'insumos'
    """
    if tipo not in ['equipos', 'insumos']:
        messages.error(request, 'Tipo de datos no válido')
        return redirect('login:dashboard')
    
    context = {
        'tipo': tipo
    }
    
    return render(request, 'seleccionar_metodo_carga.html', context)


@login_required
def importar_excel_view(request, tipo):
    """
    Vista para importar datos desde Excel/CSV
    """
    if request.method == 'POST':
        return procesar_importacion_excel(request, tipo)
    
    context = {
        'tipo': tipo,
        'formatos_soportados': ['.xlsx', '.xls', '.csv'],
        'tamano_maximo': '10 MB'
    }
    
    return render(request, 'importar_excel.html', context)


def procesar_importacion_excel(request, tipo):
    """
    Procesa la importación de archivos Excel/CSV
    """
    try:
        # Verificar que se subió un archivo
        if 'archivo' not in request.FILES:
            messages.error(request, 'No se seleccionó ningún archivo')
            return redirect('core:importar_excel', tipo=tipo)
        
        archivo = request.FILES['archivo']
        
        # Validar tipo de archivo
        nombre_archivo = archivo.name.lower()
        if not (nombre_archivo.endswith('.xlsx') or 
                nombre_archivo.endswith('.xls') or 
                nombre_archivo.endswith('.csv')):
            messages.error(request, 'Formato de archivo no soportado. Use .xlsx, .xls o .csv')
            return redirect('core:importar_excel', tipo=tipo)
        
        # Validar tamaño (10 MB máximo)
        if archivo.size > 10 * 1024 * 1024:
            messages.error(request, 'El archivo es demasiado grande. Máximo 10 MB')
            return redirect('core:importar_excel', tipo=tipo)
        
        # Procesar según el tipo
        if tipo == 'equipos':
            resultado = procesar_excel_equipos(archivo)
        elif tipo == 'insumos':
            resultado = procesar_excel_insumos(archivo)
        else:
            messages.error(request, 'Tipo de importación no válido')
            return redirect('core:seleccionar_metodo', tipo=tipo)
        
        # Mostrar resultados
        if resultado['exitoso']:
            messages.success(request, 
                f'Importación exitosa: {resultado["creados"]} registros creados, '
                f'{resultado["actualizados"]} actualizados, '
                f'{resultado["errores"]} errores')
        else:
            messages.error(request, f'Error en la importación: {resultado["mensaje"]}')
        
        return redirect('core:resultado_importacion', tipo=tipo)
        
    except Exception as e:
        messages.error(request, f'Error inesperado durante la importación: {str(e)}')
        return redirect('core:importar_excel', tipo=tipo)


def procesar_excel_equipos(archivo):
    """
    Procesa la importación específica de equipos
    """
    try:
        # Leer el archivo Excel/CSV
        if archivo.name.lower().endswith('.csv'):
            df = pd.read_csv(archivo, encoding='utf-8')
        else:
            df = pd.read_excel(archivo)
        
        # Validar columnas requeridas para equipos
        columnas_requeridas = [
            'Unidad Académica', 'Carrera', 'Semestre', 'Asignatura',
            'Nombre de Equipo Existente', 'Estado'
        ]
        
        columnas_faltantes = [col for col in columnas_requeridas if col not in df.columns]
        if columnas_faltantes:
            return {
                'exitoso': False,
                'mensaje': f'Faltan columnas requeridas: {", ".join(columnas_faltantes)}'
            }
        
        # Procesar cada fila
        creados = 0
        actualizados = 0
        errores = 0
        
        for index, row in df.iterrows():
            try:
                # Aquí iría la lógica específica de creación de equipos
                # Por ahora simulo el proceso
                creados += 1
            except Exception as e:
                errores += 1
                print(f"Error en fila {index + 2}: {str(e)}")
        
        return {
            'exitoso': True,
            'creados': creados,
            'actualizados': actualizados,
            'errores': errores
        }
        
    except Exception as e:
        return {
            'exitoso': False,
            'mensaje': str(e)
        }


def procesar_excel_insumos(archivo):
    """
    Procesa la importación específica de insumos
    """
    try:
        # Leer el archivo Excel/CSV
        if archivo.name.lower().endswith('.csv'):
            df = pd.read_csv(archivo, encoding='utf-8')
        else:
            df = pd.read_excel(archivo)
        
        # Validar columnas requeridas para insumos
        columnas_requeridas = [
            'Unidad Académica', 'Laboratorio', 'Categoría',
            'Nombre del Elemento', 'Descripción/Características'
        ]
        
        columnas_faltantes = [col for col in columnas_requeridas if col not in df.columns]
        if columnas_faltantes:
            return {
                'exitoso': False,
                'mensaje': f'Faltan columnas requeridas: {", ".join(columnas_faltantes)}'
            }
        
        # Procesar cada fila
        creados = 0
        actualizados = 0
        errores = 0
        
        for index, row in df.iterrows():
            try:
                # Aquí iría la lógica específica de creación de insumos
                # Por ahora simulo el proceso
                creados += 1
            except Exception as e:
                errores += 1
                print(f"Error en fila {index + 2}: {str(e)}")
        
        return {
            'exitoso': True,
            'creados': creados,
            'actualizados': actualizados,
            'errores': errores
        }
        
    except Exception as e:
        return {
            'exitoso': False,
            'mensaje': str(e)
        }


@login_required
def descargar_plantilla_excel(request, tipo):
    """
    Genera y descarga la plantilla oficial de Excel para el tipo especificado
    """
    try:
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        
        if tipo == 'equipos':
            ws.title = "Plantilla Equipos EMI"
            
            # Encabezados para las 24 columnas oficiales
            headers = [
                'Unidad Académica', 'Carrera', 'Semestre', 'Asignatura',
                'Carga Horaria Semanal', 'Carga Horaria Semestral',
                'Criterio de Desempeño', 'Unidad Didáctica', 'Contenido Analítico',
                'Guía de Laboratorio', 'Práctica', 'Nombre de Equipo Existente',
                'Marca', 'Modelo', 'Estado', 'Número de Unidades del Equipo',
                'Es un Activo Fijo', 'Fotografía Frontal', 'Fotografía de Placa',
                'Ubicación del Equipo (Laboratorio)', 'Sección/Área',
                'Identificador/Nº de Aula', 'Equipo Requerido', 'Número de Equipos Requeridos'
            ]
            
        elif tipo == 'insumos':
            ws.title = "Plantilla Insumos EMI"
            
            # Encabezados para las 19 columnas oficiales de insumos
            headers = [
                'Unidad Académica', 'Laboratorio', 'Categoría', 'Nombre del Elemento',
                'Descripción/Características', 'Marca/Modelo', 'Número de Serie',
                'Código de Barras', 'Proveedor', 'Fecha de Adquisición',
                'Fecha de Vencimiento', 'Cantidad/Stock', 'Unidad de Medida',
                'Costo Unitario', 'Uso Principal', 'Condiciones de Almacenamiento',
                'Observaciones', 'Estado', 'Responsable'
            ]
        
        # Escribir encabezados
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)
        
        # Agregar filas de ejemplo
        if tipo == 'equipos':
            ejemplo = [
                'UALP', 'ING_SISTEMAS', '3', 'programacion_estructurada',
                '4', '80', 'Criterio ejemplo', 'Unidad 1', 'Contenido ejemplo',
                'Guía 1', 'Práctica 1', 'Computadora Desktop',
                'Dell', 'OptiPlex 3080', 'bueno', '1',
                'No', '', '', 'LAB_SISTEMAS_1', 'Área A',
                'Aula 101', 'Computadora', '25'
            ]
        elif tipo == 'insumos':
            ejemplo = [
                'UALP', 'LAB_QUIMICA', 'reactivos', 'Ácido Clorhídrico',
                'Solución al 37% grado analítico', 'Merck', 'HCL001',
                '7501234567890', 'Distribuidora Química SA', '2024-01-15',
                '2025-01-15', '500', 'ml', '25.50',
                'ensayos', 'lugar_ventilado', 'Usar con precaución', 'bueno', 'Dr. Juan Pérez'
            ]
        
        # Escribir fila de ejemplo
        for col, valor in enumerate(ejemplo, 1):
            ws.cell(row=2, column=col, value=valor)
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Preparar respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="Plantilla_{tipo.title()}_EMI.xlsx"'
        
        # Guardar el workbook en la respuesta
        wb.save(response)
        
        return response
        
    except Exception as e:
        messages.error(request, f'Error al generar la plantilla: {str(e)}')
        return redirect('core:seleccionar_metodo', tipo=tipo)