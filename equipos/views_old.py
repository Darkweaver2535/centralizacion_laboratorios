from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.db import transaction
from django.contrib import messages
from django.db.models import Q, Count
from django.core.paginator import Paginator
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.utils import timezone
import json

from .models import Equipo, HistorialEquipo, MantenimientoEquipo
from core.models import UnidadAcademica, Carrera, Asignatura, UnidadTematica, GuiaLaboratorio, Practica, Laboratorio

@login_required
def equipos_view(request):
    """Vista principal de equipos con filtros y paginación"""
    
    # Obtener filtros
    filtros = {
        'unidad_academica': request.GET.get('unidad_academica', ''),
        'carrera': request.GET.get('carrera', ''),
        'semestre': request.GET.get('semestre', ''),
        'asignatura': request.GET.get('asignatura', ''),
        'estado': request.GET.get('estado', ''),
        'laboratorio': request.GET.get('laboratorio', ''),
        'busqueda': request.GET.get('busqueda', ''),
    }
    
    # Construir queryset con filtros
    equipos = Equipo.objects.all()
    
    if filtros['unidad_academica']:
        equipos = equipos.filter(unidad_academica_id=filtros['unidad_academica'])
    
    if filtros['carrera']:
        equipos = equipos.filter(carrera_id=filtros['carrera'])
    
    if filtros['semestre']:
        equipos = equipos.filter(semestre=filtros['semestre'])
    
    if filtros['asignatura']:
        equipos = equipos.filter(asignatura_id=filtros['asignatura'])
    
    if filtros['estado']:
        equipos = equipos.filter(estado=filtros['estado'])
    
    if filtros['laboratorio']:
        equipos = equipos.filter(laboratorio_id=filtros['laboratorio'])
    
    if filtros['busqueda']:
        equipos = equipos.filter(
            Q(equipo_existente__icontains=filtros['busqueda']) |
            Q(marca__icontains=filtros['busqueda']) |
            Q(modelo__icontains=filtros['busqueda']) |
            Q(codigo_inventario__icontains=filtros['busqueda'])
        )
    
    # Ordenar por fecha de creación (más recientes primero)
    equipos = equipos.order_by('-created_at')
    
    # Paginación
    paginator = Paginator(equipos, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Datos para los filtros
    unidades = UnidadAcademica.objects.all()
    carreras = Carrera.objects.all()
    laboratorios = Laboratorio.objects.all()
    
    # Estadísticas
    stats = {
        'total': equipos.count(),
        'operativos': equipos.filter(estado='operativo').count(),
        'mantenimiento': equipos.filter(estado='mantenimiento').count(),
        'reparacion': equipos.filter(estado='reparacion').count(),
        'inoperativos': equipos.filter(estado='inoperativo').count(),
    }
    
    context = {
        'page_obj': page_obj,
        'filtros': filtros,
        'unidades': unidades,
        'carreras': carreras,
        'laboratorios': laboratorios,
        'stats': stats,
        'estados_choices': Equipo.ESTADOS,
        'semestres_choices': [(i, f"{i}° Semestre") for i in range(1, 11)],
    }
    
    return render(request, 'equipos/lista.html', context)

@login_required
def nuevo_equipo_view(request):
    """Vista para crear un nuevo equipo"""
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Crear el equipo con todos los datos del formulario
                equipo = Equipo.objects.create(
                    unidad_academica_id=request.POST.get('unidad_academica'),
                    carrera_id=request.POST.get('carrera'),
                    semestre=int(request.POST.get('semestre')),
                    asignatura_id=request.POST.get('asignatura'),
                    carga_horaria_semanal=int(request.POST.get('carga_horaria_semanal')),
                    carga_horaria_semestral=int(request.POST.get('carga_horaria_semestral')),
                    unidad_tematica_id=request.POST.get('unidad_tematica'),
                    guia_laboratorio_id=request.POST.get('guia_laboratorio'),
                    practica_id=request.POST.get('practica'),
                    equipo_existente=request.POST.get('equipo_existente'),
                    marca=request.POST.get('marca', ''),
                    modelo=request.POST.get('modelo', ''),
                    estado=request.POST.get('estado', 'operativo'),
                    numero_unidades=int(request.POST.get('numero_unidades', 1)),
                    es_activo_fijo=request.POST.get('es_activo_fijo') == 'on',
                    laboratorio_id=request.POST.get('laboratorio'),
                    seccion_area=request.POST.get('seccion_area', ''),
                    identificador_aula=request.POST.get('identificador_aula', ''),
                    equipo_requerido=request.POST.get('equipo_requerido', ''),
                    numero_equipos_requeridos=int(request.POST.get('numero_equipos_requeridos', 0)),
                    usuario_creador=request.user,
                    observaciones=request.POST.get('observaciones', ''),
                )
                
                # Manejar archivos de imagen
                if 'fotografia_frontal' in request.FILES:
                    equipo.fotografia_frontal = request.FILES['fotografia_frontal']
                
                if 'fotografia_placa' in request.FILES:
                    equipo.fotografia_placa = request.FILES['fotografia_placa']
                
                equipo.save()
                
                messages.success(request, f'Equipo "{equipo.equipo_existente}" creado exitosamente.')
                return redirect('equipos:detalle', pk=equipo.pk)
                
        except Exception as e:
            messages.error(request, f'Error al crear el equipo: {str(e)}')
    
    # Datos para los formularios
    unidades = UnidadAcademica.objects.all()
    carreras = Carrera.objects.all()
    laboratorios = Laboratorio.objects.all()
    
    context = {
        'unidades': unidades,
        'carreras': carreras,
        'laboratorios': laboratorios,
        'estados_choices': Equipo.ESTADOS,
        'semestres_choices': [(i, f"{i}° Semestre") for i in range(1, 11)],
    }
    
    return render(request, 'equipos/nuevo.html', context)

@login_required
def detalle_equipo_view(request, pk):
    """Vista detalle de un equipo específico"""
    equipo = get_object_or_404(Equipo, pk=pk)
    
    # Obtener historial de cambios
    historial = HistorialEquipo.objects.filter(equipo=equipo).order_by('-fecha_cambio')[:10]
    
    # Obtener mantenimientos
    mantenimientos = MantenimientoEquipo.objects.filter(equipo=equipo).order_by('-fecha_inicio')[:10]
    
    context = {
        'equipo': equipo,
        'historial': historial,
        'mantenimientos': mantenimientos,
    }
    
    return render(request, 'equipos/detalle.html', context)

@login_required
def editar_equipo_view(request, pk):
    """Vista para editar un equipo existente"""
    equipo = get_object_or_404(Equipo, pk=pk)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Registrar cambio de estado si aplica
                estado_anterior = equipo.estado
                estado_nuevo = request.POST.get('estado', equipo.estado)
                
                if estado_anterior != estado_nuevo:
                    HistorialEquipo.objects.create(
                        equipo=equipo,
                        estado_anterior=estado_anterior,
                        estado_nuevo=estado_nuevo,
                        usuario=request.user,
                        observaciones=f"Cambio de estado: {estado_anterior} → {estado_nuevo}"
                    )
                
                # Actualizar todos los campos
                equipo.unidad_academica_id = request.POST.get('unidad_academica')
                equipo.carrera_id = request.POST.get('carrera')
                equipo.semestre = int(request.POST.get('semestre'))
                equipo.asignatura_id = request.POST.get('asignatura')
                equipo.carga_horaria_semanal = int(request.POST.get('carga_horaria_semanal'))
                equipo.carga_horaria_semestral = int(request.POST.get('carga_horaria_semestral'))
                equipo.unidad_tematica_id = request.POST.get('unidad_tematica')
                equipo.guia_laboratorio_id = request.POST.get('guia_laboratorio')
                equipo.practica_id = request.POST.get('practica')
                equipo.equipo_existente = request.POST.get('equipo_existente')
                equipo.marca = request.POST.get('marca', '')
                equipo.modelo = request.POST.get('modelo', '')
                equipo.estado = estado_nuevo
                equipo.numero_unidades = int(request.POST.get('numero_unidades', 1))
                equipo.es_activo_fijo = request.POST.get('es_activo_fijo') == 'on'
                equipo.laboratorio_id = request.POST.get('laboratorio')
                equipo.seccion_area = request.POST.get('seccion_area', '')
                equipo.identificador_aula = request.POST.get('identificador_aula', '')
                equipo.equipo_requerido = request.POST.get('equipo_requerido', '')
                equipo.numero_equipos_requeridos = int(request.POST.get('numero_equipos_requeridos', 0))
                equipo.observaciones = request.POST.get('observaciones', '')
                
                # Manejar archivos de imagen
                if 'fotografia_frontal' in request.FILES:
                    equipo.fotografia_frontal = request.FILES['fotografia_frontal']
                
                if 'fotografia_placa' in request.FILES:
                    equipo.fotografia_placa = request.FILES['fotografia_placa']
                
                equipo.save()
                
                messages.success(request, f'Equipo "{equipo.equipo_existente}" actualizado exitosamente.')
                return redirect('equipos:detalle', pk=equipo.pk)
                
        except Exception as e:
            messages.error(request, f'Error al actualizar el equipo: {str(e)}')
    
    # Datos para los formularios
    unidades = UnidadAcademica.objects.all()
    carreras = Carrera.objects.all()
    laboratorios = Laboratorio.objects.all()
    
    # Datos relacionados del equipo actual
    asignaturas = Asignatura.objects.filter(carrera=equipo.carrera, semestre=equipo.semestre) if equipo.carrera else []
    unidades_tematicas = UnidadTematica.objects.filter(asignatura=equipo.asignatura) if equipo.asignatura else []
    guias_laboratorio = GuiaLaboratorio.objects.filter(unidad_tematica=equipo.unidad_tematica) if equipo.unidad_tematica else []
    practicas = Practica.objects.filter(guia_laboratorio=equipo.guia_laboratorio) if equipo.guia_laboratorio else []
    
    context = {
        'equipo': equipo,
        'unidades': unidades,
        'carreras': carreras,
        'laboratorios': laboratorios,
        'asignaturas': asignaturas,
        'unidades_tematicas': unidades_tematicas,
        'guias_laboratorio': guias_laboratorio,
        'practicas': practicas,
        'estados_choices': Equipo.ESTADOS,
        'semestres_choices': [(i, f"{i}° Semestre") for i in range(1, 11)],
    }
    
    return render(request, 'equipos/editar.html', context)

@login_required
def eliminar_equipo_view(request, pk):
    """Vista para eliminar un equipo"""
    equipo = get_object_or_404(Equipo, pk=pk)
    
    if request.method == 'POST':
        nombre_equipo = equipo.equipo_existente
        equipo.delete()
        messages.success(request, f'Equipo "{nombre_equipo}" eliminado exitosamente.')
        return redirect('equipos:lista')
    
    context = {'equipo': equipo}
    return render(request, 'equipos/eliminar.html', context)
def ingreso_equipos_view(request):
    """Vista para el ingreso de equipos"""
    if request.method == 'POST':
# AJAX Views para carga dinámica de datos relacionados

@login_required
def get_carreras_ajax(request):
    """Obtener carreras por unidad académica"""
    unidad_id = request.GET.get('unidad_id')
    
    if unidad_id:
        # Filtrar carreras que tienen equipos en esa unidad académica
        carreras = Carrera.objects.filter(
            id__in=Equipo.objects.filter(unidad_academica_id=unidad_id).values_list('carrera_id', flat=True).distinct()
        ).distinct()
    else:
        carreras = Carrera.objects.all()
    
    carreras_data = [
        {'id': carrera.id, 'nombre': carrera.nombre, 'display': carrera.get_nombre_display()}
        for carrera in carreras
    ]
    
    return JsonResponse({'carreras': carreras_data})

@login_required
def get_asignaturas_ajax(request):
    """Obtener asignaturas por carrera y semestre"""
    carrera_id = request.GET.get('carrera_id')
    semestre = request.GET.get('semestre')
    
    asignaturas = Asignatura.objects.all()
    
    if carrera_id:
        asignaturas = asignaturas.filter(carrera_id=carrera_id)
    
    if semestre:
        asignaturas = asignaturas.filter(semestre=semestre)
    
    asignaturas_data = [
        {
            'id': asignatura.id, 
            'nombre': asignatura.nombre, 
            'display': asignatura.get_nombre_display(),
            'carga_semanal': asignatura.carga_horaria_semanal,
            'carga_semestral': asignatura.carga_horaria_semestral
        }
        for asignatura in asignaturas
    ]
    
    return JsonResponse({'asignaturas': asignaturas_data})

@login_required
def get_unidades_tematicas_ajax(request):
    """Obtener unidades temáticas por asignatura"""
    asignatura_id = request.GET.get('asignatura_id')
    
    if asignatura_id:
        unidades = UnidadTematica.objects.filter(asignatura_id=asignatura_id)
    else:
        unidades = UnidadTematica.objects.none()
    
    unidades_data = [
        {'id': unidad.id, 'numero': unidad.numero, 'nombre': unidad.nombre}
        for unidad in unidades
    ]
    
    return JsonResponse({'unidades_tematicas': unidades_data})

@login_required
def get_guias_laboratorio_ajax(request):
    """Obtener guías de laboratorio por unidad temática"""
    unidad_tematica_id = request.GET.get('unidad_tematica_id')
    
    if unidad_tematica_id:
        guias = GuiaLaboratorio.objects.filter(unidad_tematica_id=unidad_tematica_id)
    else:
        guias = GuiaLaboratorio.objects.none()
    
    guias_data = [
        {'id': guia.id, 'numero': guia.numero, 'nombre': guia.nombre}
        for guia in guias
    ]
    
    return JsonResponse({'guias_laboratorio': guias_data})

@login_required
def get_practicas_ajax(request):
    """Obtener prácticas por guía de laboratorio"""
    guia_id = request.GET.get('guia_id')
    
    if guia_id:
        practicas = Practica.objects.filter(guia_laboratorio_id=guia_id)
    else:
        practicas = Practica.objects.none()
    
    practicas_data = [
        {'id': practica.id, 'numero': practica.numero, 'nombre': practica.nombre}
        for practica in practicas
    ]
    
    return JsonResponse({'practicas': practicas_data})

@login_required
def get_laboratorios_ajax(request):
    """Obtener laboratorios por unidad académica"""
    unidad_id = request.GET.get('unidad_id')
    
    if unidad_id:
        laboratorios = Laboratorio.objects.filter(unidad_academica_id=unidad_id)
    else:
        laboratorios = Laboratorio.objects.all()
    
    laboratorios_data = [
        {
            'id': laboratorio.id, 
            'nombre': laboratorio.nombre, 
            'display': laboratorio.get_nombre_display(),
            'seccion_area': laboratorio.seccion_area,
            'identificador_aula': laboratorio.identificador_aula
        }
        for laboratorio in laboratorios
    ]
    
    return JsonResponse({'laboratorios': laboratorios_data})

@login_required
def filtrar_equipos_ajax(request):
    """Filtrar equipos via AJAX para la vista de lista"""
    
    # Obtener filtros del request
    filtros = {
        'unidad_academica': request.GET.get('unidad_academica', ''),
        'carrera': request.GET.get('carrera', ''),
        'semestre': request.GET.get('semestre', ''),
        'asignatura': request.GET.get('asignatura', ''),
        'estado': request.GET.get('estado', ''),
        'laboratorio': request.GET.get('laboratorio', ''),
        'busqueda': request.GET.get('busqueda', ''),
    }
    
    # Construir queryset
    equipos = Equipo.objects.all()
    
    if filtros['unidad_academica']:
        equipos = equipos.filter(unidad_academica_id=filtros['unidad_academica'])
    
    if filtros['carrera']:
        equipos = equipos.filter(carrera_id=filtros['carrera'])
    
    if filtros['semestre']:
        equipos = equipos.filter(semestre=filtros['semestre'])
    
    if filtros['asignatura']:
        equipos = equipos.filter(asignatura_id=filtros['asignatura'])
    
    if filtros['estado']:
        equipos = equipos.filter(estado=filtros['estado'])
    
    if filtros['laboratorio']:
        equipos = equipos.filter(laboratorio_id=filtros['laboratorio'])
    
    if filtros['busqueda']:
        equipos = equipos.filter(
            Q(equipo_existente__icontains=filtros['busqueda']) |
            Q(marca__icontains=filtros['busqueda']) |
            Q(modelo__icontains=filtros['busqueda']) |
            Q(codigo_inventario__icontains=filtros['busqueda'])
        )
    
    # Preparar datos para respuesta
    equipos_data = []
    for equipo in equipos[:50]:  # Limitar a 50 resultados
        equipos_data.append({
            'id': equipo.id,
            'codigo_inventario': equipo.codigo_inventario,
            'equipo_existente': equipo.equipo_existente,
            'marca': equipo.marca,
            'modelo': equipo.modelo,
            'estado': equipo.get_estado_display(),
            'unidad_academica': equipo.unidad_academica.get_nombre_display(),
            'carrera': equipo.carrera.get_nombre_display(),
            'semestre': f"{equipo.semestre}° Semestre",
            'asignatura': equipo.asignatura.get_nombre_display(),
            'laboratorio': equipo.laboratorio.get_nombre_display(),
            'numero_unidades': equipo.numero_unidades,
            'created_at': equipo.created_at.strftime('%d/%m/%Y'),
        })
    
    # Estadísticas
    stats = {
        'total': equipos.count(),
        'operativos': equipos.filter(estado='operativo').count(),
        'mantenimiento': equipos.filter(estado='mantenimiento').count(),
        'reparacion': equipos.filter(estado='reparacion').count(),
        'inoperativos': equipos.filter(estado='inoperativo').count(),
    }
    
    return JsonResponse({
        'equipos': equipos_data,
        'stats': stats,
        'total_equipos': equipos.count()
    })

@login_required
def exportar_equipos_excel(request):
    """Exportar equipos a Excel con todas las 22 columnas"""
    
    # Obtener filtros si existen
    filtros = {
        'unidad_academica': request.GET.get('unidad_academica', ''),
        'carrera': request.GET.get('carrera', ''),
        'semestre': request.GET.get('semestre', ''),
        'estado': request.GET.get('estado', ''),
    }
    
    # Construir queryset
    equipos = Equipo.objects.select_related(
        'unidad_academica', 'carrera', 'asignatura', 'unidad_tematica',
        'guia_laboratorio', 'practica', 'laboratorio', 'usuario_creador'
    ).all()
    
    # Aplicar filtros
    if filtros['unidad_academica']:
        equipos = equipos.filter(unidad_academica_id=filtros['unidad_academica'])
    
    if filtros['carrera']:
        equipos = equipos.filter(carrera_id=filtros['carrera'])
    
    if filtros['semestre']:
        equipos = equipos.filter(semestre=filtros['semestre'])
    
    if filtros['estado']:
        equipos = equipos.filter(estado=filtros['estado'])
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Equipos de Laboratorio"
    
    # Definir estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )
    
    # Encabezados (22 columnas)
    headers = [
        'UNIDAD ACADÉMICA',
        'CARRERA',
        'SEMESTRE',
        'ASIGNATURA',
        'CARGA HORARIA SEMANAL',
        'CARGA HORARIA SEMESTRAL',
        'UNIDAD TEMÁTICA',
        'GUÍA DE LABORATORIO',
        'PRÁCTICA',
        'EQUIPO EXISTENTE',
        'MARCA',
        'MODELO',
        'ESTADO',
        'NÚMERO DE UNIDADES DEL EQUIPO',
        'ES UN ACTIVO FIJO DE ACUERDO A SU ACTA DE ENTREGA?',
        'FOTOGRAFÍA FRONTAL DEL EQUIPO',
        'FOTOGRAFÍA DE LA PLACA DE CARACTERÍSTICAS',
        'UBICACIÓN DEL EQUIPO (LABORATORIO)',
        'SECCIÓN/ÁREA',
        'IDENTIFICADOR/Nº DE AULA',
        'EQUIPO REQUERIDO',
        'NÚMERO DE EQUIPOS REQUERIDOS'
    ]
    
    # Escribir encabezados
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Escribir datos
    for row, equipo in enumerate(equipos, 2):
        data = [
            equipo.unidad_academica.get_nombre_display(),
            equipo.carrera.get_nombre_display(),
            f"{equipo.semestre}° Semestre",
            equipo.asignatura.get_nombre_display(),
            equipo.carga_horaria_semanal,
            equipo.carga_horaria_semestral,
            f"Unidad {equipo.unidad_tematica.numero}: {equipo.unidad_tematica.nombre}",
            f"Guía {equipo.guia_laboratorio.numero}: {equipo.guia_laboratorio.nombre}",
            f"Práctica {equipo.practica.numero}: {equipo.practica.nombre}",
            equipo.equipo_existente,
            equipo.marca,
            equipo.modelo,
            equipo.get_estado_display(),
            equipo.numero_unidades,
            'Sí' if equipo.es_activo_fijo else 'No',
            'Sí' if equipo.fotografia_frontal else 'No',
            'Sí' if equipo.fotografia_placa else 'No',
            equipo.laboratorio.get_nombre_display(),
            equipo.seccion_area,
            equipo.identificador_aula,
            equipo.equipo_requerido,
            equipo.numero_equipos_requeridos,
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
    
    # Ajustar ancho de columnas
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 20
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    fecha_actual = timezone.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="equipos_laboratorio_{fecha_actual}.xlsx"'
    
    wb.save(response)
    return response
    
    return render(request, 'equipos/ingreso_equipos.html', context)

@login_required
def visualizacion_equipos_view(request):
    """Vista para la visualización de equipos"""
    
    # Obtener datos para filtros
    unidades = UnidadAcademica.objects.all()
    carreras = Carrera.objects.all()
    tipos_equipos = TipoEquipo.objects.all()
    laboratorios = Laboratorio.objects.all()
    
    # Estadísticas básicas
    stats = {
        'total_equipos_existentes': EquipoExistente.objects.count(),
        'total_equipos_requeridos': EquipoRequerido.objects.count(),
        'total_tipos_equipos': TipoEquipo.objects.count(),
        'total_laboratorios': Laboratorio.objects.count(),
    }
    
    context = {
        'unidades': unidades,
        'carreras': carreras,
        'tipos_equipos': tipos_equipos,
        'laboratorios': laboratorios,
        'stats': stats,
    }
    
    return render(request, 'equipos/visualizacion_equipos.html', context)

@csrf_exempt
def obtener_asignaturas_ajax(request):
    """Vista AJAX para obtener asignaturas por carrera y semestre"""
    if request.method == 'POST':
        carrera_id = request.POST.get('carrera')
        semestre = request.POST.get('semestre')
        
        if carrera_id and semestre:
            try:
                carrera = Carrera.objects.get(nombre=carrera_id)
                semestre_int = int(semestre)
                
                # Obtener asignaturas para el semestre
                asignaturas = Asignatura.get_asignaturas_por_semestre(semestre_int)
                
                return JsonResponse({
                    'asignaturas': asignaturas
                })
            except (Carrera.DoesNotExist, ValueError):
                pass
    
    return JsonResponse({'asignaturas': []})

@csrf_exempt
def obtener_unidades_tematicas_ajax(request):
    """Vista AJAX para obtener unidades temáticas por asignatura"""
    if request.method == 'POST':
        asignatura_id = request.POST.get('asignatura')
        
        if asignatura_id:
            try:
                asignatura = Asignatura.objects.get(nombre=asignatura_id)
                unidades = UnidadTematica.objects.filter(asignatura=asignatura)
                
                unidades_data = []
                for unidad in unidades:
                    unidades_data.append({
                        'id': unidad.id,
                        'nombre': unidad.nombre,
                        'numero': unidad.numero
                    })
                
                return JsonResponse({
                    'unidades_tematicas': unidades_data
                })
            except Asignatura.DoesNotExist:
                pass
    
    return JsonResponse({'unidades_tematicas': []})

@csrf_exempt
def obtener_guias_laboratorio_ajax(request):
    """Vista AJAX para obtener guías de laboratorio por unidad temática"""
    if request.method == 'POST':
        unidad_tematica_id = request.POST.get('unidad_tematica')
        
        if unidad_tematica_id:
            try:
                unidad = UnidadTematica.objects.get(id=unidad_tematica_id)
                guias = GuiaLaboratorio.objects.filter(unidad_tematica=unidad)
                
                guias_data = []
                for guia in guias:
                    guias_data.append({
                        'id': guia.id,
                        'nombre': guia.nombre,
                        'numero': guia.numero
                    })
                
                return JsonResponse({
                    'guias_laboratorio': guias_data
                })
            except UnidadTematica.DoesNotExist:
                pass
    
    return JsonResponse({'guias_laboratorio': []})

@csrf_exempt
def obtener_practicas_ajax(request):
    """Vista AJAX para obtener prácticas por guía de laboratorio"""
    if request.method == 'POST':
        guia_id = request.POST.get('guia_laboratorio')
        
        if guia_id:
            try:
                guia = GuiaLaboratorio.objects.get(id=guia_id)
                practicas = Practica.objects.filter(guia_laboratorio=guia)
                
                practicas_data = []
                for practica in practicas:
                    practicas_data.append({
                        'id': practica.id,
                        'nombre': practica.nombre,
                        'numero': practica.numero
                    })
                
                return JsonResponse({
                    'practicas': practicas_data
                })
            except GuiaLaboratorio.DoesNotExist:
                pass
    
    return JsonResponse({'practicas': []})

@csrf_exempt
def obtener_laboratorios_ajax(request):
    """Vista AJAX para obtener laboratorios por unidad académica"""
    if request.method == 'POST':
        unidad_academica_id = request.POST.get('unidad_academica')
        
        if unidad_academica_id:
            try:
                unidad = UnidadAcademica.objects.get(nombre=unidad_academica_id)
                laboratorios = Laboratorio.objects.filter(unidad_academica=unidad)
                
                laboratorios_data = []
                for lab in laboratorios:
                    laboratorios_data.append({
                        'nombre': lab.nombre,
                        'display_name': lab.get_nombre_display(),
                        'seccion_area': lab.seccion_area,
                        'identificador_aula': lab.identificador_aula
                    })
                
                return JsonResponse({
                    'laboratorios': laboratorios_data
                })
            except UnidadAcademica.DoesNotExist:
                pass
    
    return JsonResponse({'laboratorios': []})

@login_required
def filtrar_equipos(request):
    """Vista para filtrar equipos según criterios"""
    filtros = {
        'unidad_academica': request.GET.get('unidad_academica'),
        'carrera': request.GET.get('carrera'),
        'asignatura': request.GET.get('asignatura'),
        'laboratorio': request.GET.get('laboratorio'),
        'tipo_equipo': request.GET.get('tipo_equipo'),
    }
    
    # Filtrar equipos existentes
    equipos_existentes = EquipoExistente.objects.all()
    
    if filtros['unidad_academica']:
        equipos_existentes = equipos_existentes.filter(
            laboratorio__unidad_academica__nombre=filtros['unidad_academica']
        )
    
    if filtros['tipo_equipo']:
        equipos_existentes = equipos_existentes.filter(
            tipo_equipo__nombre=filtros['tipo_equipo']
        )
    
    # Preparar datos para respuesta
    equipos_data = []
    for equipo in equipos_existentes:
        equipos_data.append({
            'id': equipo.id,
            'tipo_equipo': equipo.tipo_equipo.get_nombre_display(),
            'marca': equipo.marca,
            'modelo': equipo.modelo,
            'estado': equipo.get_estado_display(),
            'numero_unidades': equipo.numero_unidades,
            'es_activo_fijo': equipo.es_activo_fijo,
            'laboratorio': equipo.laboratorio.get_nombre_display(),
            'unidad_academica': equipo.laboratorio.unidad_academica.get_nombre_display(),
        })
    
    return JsonResponse({
        'equipos': equipos_data,
        'total': len(equipos_data)
    })

@login_required
def exportar_equipos_excel(request):
    """Vista para exportar equipos a Excel"""
    # Implementar lógica de exportación similar a visualización
    from django.http import HttpResponse
    import openpyxl
    from datetime import datetime
    
    # Obtener filtros
    filtros = {
        'unidad_academica': request.GET.get('unidad_academica'),
        'carrera': request.GET.get('carrera'),
        'tipo_equipo': request.GET.get('tipo_equipo'),
    }
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Equipos"
    
    # Headers según las nuevas columnas
    headers = [
        'UNIDAD ACADÉMICA', 'CARRERA', 'SEMESTRE', 'ASIGNATURA', 
        'CARGA HORARIA SEMANAL', 'CARGA HORARIA SEMESTRAL', 'UNIDAD TEMÁTICA',
        'GUÍA DE LABORATORIO', 'PRÁCTICA', 'EQUIPO EXISTENTE', 'MARCA',
        'MODELO', 'ESTADO', 'NÚMERO DE UNIDADES DEL EQUIPO',
        'ES UN ACTIVO FIJO DE ACUERDO A SU ACTA DE ENTREGA?',
        'FOTOGRAFÍA FRONTAL DEL EQUIPO', 'FOTOGRAFÍA DE LA PLACA DE CARACTERÍSTICAS',
        'UBICACIÓN DEL EQUIPO (LABORATORIO)', 'SECCIÓN/ÁREA', 'IDENTIFICADOR/Nº DE AULA',
        'EQUIPO REQUERIDO', 'NÚMERO DE EQUIPOS REQUERIDOS'
    ]
    
    # Escribir headers
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Aquí implementarías la lógica para llenar los datos
    # Por ahora, solo crear el archivo vacío
    
    # Configurar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="equipos_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    
    wb.save(response)
    return response
