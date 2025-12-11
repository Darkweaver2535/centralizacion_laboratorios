from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.db import transaction
from django.contrib import messages
from django.db.models import Q, Count
from django.core.paginator import Paginator
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.utils import timezone
import json

from .models import Equipo, HistorialEquipo, MantenimientoEquipo, TareaReordenamiento, EquipoTarea, LogReordenamiento
from .forms import EquipoForm
from core.models import (
    UnidadAcademica, Carrera, Asignatura, UnidadTematica, 
    GuiaLaboratorio, Practica, Laboratorio, CriterioDesempeno,
    UnidadDidactica, ContenidoAnalitico
)

@login_required
def equipos_view(request):
    """Vista principal de equipos con filtros y paginación"""
    
    # Obtener filtros
    filtros = {
        'unidad_academica': request.GET.get('unidad_academica', ''),
        'carrera': request.GET.get('carrera', ''),
        'semestre': request.GET.get('semestre', ''),
        'asignatura': request.GET.get('asignatura', ''),
        'estado': request.GET.get('estado', ''),
        'laboratorio': request.GET.get('laboratorio', ''),
        'busqueda': request.GET.get('busqueda', ''),
    }
    
    # Construir queryset con filtros
    equipos = Equipo.objects.select_related(
        'unidad_academica', 'carrera', 'asignatura', 'laboratorio',
        'criterio_desempeno', 'unidad_didactica', 'contenido_analitico',
        'guia_laboratorio', 'practica'
    ).all()
    
    if filtros['unidad_academica']:
        equipos = equipos.filter(unidad_academica_id=filtros['unidad_academica'])
    
    if filtros['carrera']:
        equipos = equipos.filter(carrera_id=filtros['carrera'])
    
    if filtros['semestre']:
        equipos = equipos.filter(semestre=filtros['semestre'])
    
    if filtros['asignatura']:
        equipos = equipos.filter(asignatura_id=filtros['asignatura'])
    
    if filtros['estado']:
        equipos = equipos.filter(estado=filtros['estado'])
    
    if filtros['laboratorio']:
        equipos = equipos.filter(laboratorio_id=filtros['laboratorio'])
    
    if filtros['busqueda']:
        equipos = equipos.filter(
            Q(equipo_existente__icontains=filtros['busqueda']) |
            Q(marca__icontains=filtros['busqueda']) |
            Q(modelo__icontains=filtros['busqueda']) |
            Q(codigo_inventario__icontains=filtros['busqueda'])
        )
    
    # Ordenar por fecha de creación (más recientes primero)
    equipos = equipos.order_by('-created_at')
    
    # Paginación
    paginator = Paginator(equipos, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Datos para los filtros - Solo mostrar unidades académicas que tienen equipos
    unidades = UnidadAcademica.objects.filter(
        id__in=Equipo.objects.values_list('unidad_academica_id', flat=True).distinct()
    )
    carreras = Carrera.objects.all()
    laboratorios = Laboratorio.objects.all()
    
    # Estadísticas
    stats = {
        'total': equipos.count(),
        'bueno': equipos.filter(estado='bueno').count(),
        'regular': equipos.filter(estado='regular').count(),
        'malo': equipos.filter(estado='malo').count(),
    }
    
    context = {
        'page_obj': page_obj,
        'filtros': filtros,
        'unidades': unidades,
        'carreras': carreras,
        'laboratorios': laboratorios,
        'stats': stats,
        'estados_choices': Equipo.ESTADOS,
        'semestres_choices': [(i, f"{i}° Semestre") for i in range(1, 11)],
    }
    
    return render(request, 'equipos/lista.html', context)

@login_required
def equipos_excel_view(request):
    """Vista de equipos con formato del Excel - mostrar las 10 columnas del Excel"""
    
    # Obtener filtros específicos para la vista Excel
    filtros = {
        'unidad_academica': request.GET.get('unidad_academica', ''),
        'responsable': request.GET.get('responsable', ''),
        'codigo': request.GET.get('codigo', ''),
        'estado': request.GET.get('estado', ''),
        'busqueda': request.GET.get('busqueda', ''),
    }
    
    # Construir queryset con filtros
    equipos = Equipo.objects.select_related(
        'unidad_academica', 'carrera'
    ).all()
    
    # Aplicar filtros
    if filtros['unidad_academica']:
        equipos = equipos.filter(unidad_academica_id=filtros['unidad_academica'])
    
    if filtros['responsable']:
        equipos = equipos.filter(
            Q(responsable_excel__icontains=filtros['responsable'])
        )
    
    if filtros['codigo']:
        equipos = equipos.filter(
            Q(codigo_excel__icontains=filtros['codigo']) |
            Q(codigo_inventario__icontains=filtros['codigo'])
        )
    
    if filtros['estado']:
        # Mapear estados del Excel a estados del modelo
        estado_map = {
            'Bueno': 'bueno',
            'Regular': 'regular', 
            'Malo': 'malo'
        }
        estado_modelo = estado_map.get(filtros['estado'])
        if estado_modelo:
            equipos = equipos.filter(estado=estado_modelo)
    
    if filtros['busqueda']:
        equipos = equipos.filter(
            Q(descripcion_excel__icontains=filtros['busqueda']) |
            Q(oficina__icontains=filtros['busqueda']) |
            Q(cargo_responsable__icontains=filtros['busqueda']) |
            Q(equipo_existente__icontains=filtros['busqueda'])
        )
    
    # Ordenar por unidad académica y responsable
    equipos = equipos.order_by('unidad_academica__nombre', 'responsable_excel')
    
    # Paginación
    paginator = Paginator(equipos, 50)  # 50 equipos por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Estadísticas
    stats = {
        'total': equipos.count(),
        'bueno': equipos.filter(estado='bueno').count(),
        'regular': equipos.filter(estado='regular').count(),
        'malo': equipos.filter(estado='malo').count(),
    }
    
    # Obtener todas las unidades académicas para los filtros
    unidades = UnidadAcademica.objects.all().order_by('nombre')
    
    context = {
        'page_obj': page_obj,
        'filtros': filtros,
        'unidades': unidades,
        'stats': stats,
        'view_type': 'excel',
    }
    
    return render(request, 'equipos/lista_excel.html', context)

@login_required
def importar_equipos_view(request):
    """Vista para importar equipos desde Excel/CSV"""
    return render(request, 'equipos/importar.html', {
        'tipo': 'equipos'
    })

@login_required
def nuevo_equipo_view(request):
    """Vista para crear un nuevo equipo con todas las 24 columnas"""
    
    if request.method == 'POST':
        form = EquipoForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                with transaction.atomic():
                    equipo = form.save()
                    messages.success(request, f'Equipo "{equipo.equipo_existente}" creado exitosamente.')
                    return redirect('equipos:detalle', pk=equipo.pk)
            except Exception as e:
                messages.error(request, f'Error al crear el equipo: {str(e)}')
        else:
            messages.error(request, 'Por favor corrija los errores en el formulario.')
    else:
        form = EquipoForm()
    
    # Datos para los selectores dinámicos
    context = {
        'form': form,
        'unidades_academicas': UnidadAcademica.objects.all(),
        'laboratorios': Laboratorio.objects.all(),
        'estados_choices': Equipo.ESTADOS,
        'semestres_choices': [(i, f"{i}° Semestre") for i in range(1, 11)],
    }
    
    return render(request, 'equipos/nuevo.html', context)

@login_required
def detalle_equipo_view(request, pk):
    """Vista detalle de un equipo específico"""
    equipo = get_object_or_404(Equipo, pk=pk)
    
    # Obtener historial de cambios
    historial = HistorialEquipo.objects.filter(equipo=equipo).order_by('-fecha_cambio')[:10]
    
    # Obtener mantenimientos
    mantenimientos = MantenimientoEquipo.objects.filter(equipo=equipo).order_by('-fecha_inicio')[:10]
    
    # Calcular valores totales
    numero_unidades = equipo.numero_unidades or 1
    valor_total_dolares = None
    valor_total_bolivianos = None
    
    if hasattr(equipo, 'costo_dolares') and equipo.costo_dolares:
        valor_total_dolares = equipo.costo_dolares * numero_unidades
        
    if hasattr(equipo, 'costo_bolivianos') and equipo.costo_bolivianos:
        valor_total_bolivianos = equipo.costo_bolivianos * numero_unidades
    
    context = {
        'equipo': equipo,
        'historial': historial,
        'mantenimientos': mantenimientos,
        'valor_total_dolares': valor_total_dolares,
        'valor_total_bolivianos': valor_total_bolivianos,
    }
    
    return render(request, 'equipos/detalle.html', context)

@login_required
def editar_equipo_view(request, pk):
    """Vista para editar un equipo existente con datos del Excel de malla curricular"""
    equipo = get_object_or_404(Equipo, pk=pk)
    
    if request.method == 'POST':
        form = EquipoForm(request.POST, request.FILES, instance=equipo)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Registrar cambio de estado si aplica
                    estado_anterior = equipo.estado
                    estado_nuevo = form.cleaned_data['estado']
                    
                    if estado_anterior != estado_nuevo:
                        HistorialEquipo.objects.create(
                            equipo=equipo,
                            estado_anterior=estado_anterior,
                            estado_nuevo=estado_nuevo,
                            usuario=request.user,
                            observaciones=f"Cambio de estado: {estado_anterior} → {estado_nuevo}"
                        )
                    
                    # Guardar el equipo con los nuevos datos
                    form.save()
                    
                    messages.success(request, 'Equipo actualizado correctamente.')
                    return redirect('equipos:detalle', pk=equipo.pk)
                    
            except Exception as e:
                messages.error(request, f'Error al actualizar el equipo: {str(e)}')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = EquipoForm(instance=equipo)
    
    context = {
        'form': form,
        'equipo': equipo,
        'title': f'Editar Equipo: {equipo.equipo_existente}',
        # Datos para AJAX
        'criterios_data': list(CriterioDesempeno.objects.filter(
            asignatura=equipo.asignatura
        ).values('id', 'nombre', 'descripcion')) if equipo.asignatura else [],
        'unidades_didacticas_data': list(UnidadDidactica.objects.filter(
            asignatura=equipo.asignatura
        ).values('id', 'nombre', 'descripcion')) if equipo.asignatura else [],
        'contenidos_data': list(ContenidoAnalitico.objects.filter(
            unidad_didactica=equipo.unidad_didactica
        ).values('id', 'nombre', 'descripcion')) if equipo.unidad_didactica else [],
    }
    
    return render(request, 'equipos/editar.html', context)

@login_required
def eliminar_equipo_view(request, pk):
    """Vista para eliminar un equipo"""
    equipo = get_object_or_404(Equipo, pk=pk)
    
    if request.method == 'POST':
        nombre_equipo = equipo.equipo_existente
        equipo.delete()
        messages.success(request, 'Equipo eliminado correctamente.')
        return redirect('visualizacion:analisis')
    
    context = {'equipo': equipo}
    return render(request, 'equipos/eliminar.html', context)

# AJAX Views para carga dinámica de datos relacionados

@login_required
def get_carreras_ajax(request):
    """Obtener carreras por unidad académica"""
    unidad_id = request.GET.get('unidad_id')
    
    if unidad_id:
        # Filtrar carreras que tienen equipos en esa unidad académica
        carreras = Carrera.objects.filter(
            id__in=Equipo.objects.filter(unidad_academica_id=unidad_id).values_list('carrera_id', flat=True).distinct()
        ).distinct()
    else:
        carreras = Carrera.objects.all()
    
    carreras_data = [
        {'id': carrera.id, 'nombre': carrera.nombre, 'display': carrera.get_nombre_display()}
        for carrera in carreras
    ]
    
    return JsonResponse({'carreras': carreras_data})

@login_required
def get_asignaturas_ajax(request):
    """Obtener asignaturas por carrera y semestre"""
    carrera_id = request.GET.get('carrera_id')
    semestre = request.GET.get('semestre')
    
    asignaturas = Asignatura.objects.all()
    
    if carrera_id:
        asignaturas = asignaturas.filter(carrera_id=carrera_id)
    
    if semestre:
        asignaturas = asignaturas.filter(semestre=semestre)
    
    asignaturas_data = [
        {
            'id': asignatura.id, 
            'nombre': asignatura.nombre, 
            'display': asignatura.get_nombre_display(),
            'carga_semanal': asignatura.carga_horaria_semanal,
            'carga_semestral': asignatura.carga_horaria_semestral
        }
        for asignatura in asignaturas
    ]
    
    return JsonResponse({'asignaturas': asignaturas_data})

@login_required
def get_unidades_tematicas_ajax(request):
    """Obtener unidades temáticas por asignatura"""
    asignatura_id = request.GET.get('asignatura_id')
    
    if asignatura_id:
        unidades = UnidadTematica.objects.filter(asignatura_id=asignatura_id)
    else:
        unidades = UnidadTematica.objects.none()
    
    unidades_data = [
        {'id': unidad.id, 'numero': unidad.numero, 'nombre': unidad.nombre}
        for unidad in unidades
    ]
    
    return JsonResponse({'unidades_tematicas': unidades_data})

@login_required
def get_guias_laboratorio_ajax(request):
    """Obtener guías de laboratorio por asignatura"""
    asignatura_id = request.GET.get('asignatura_id')
    
    if asignatura_id:
        guias = GuiaLaboratorio.objects.filter(unidad_tematica__asignatura_id=asignatura_id)
    else:
        guias = GuiaLaboratorio.objects.none()
    
    guias_data = [
        {'id': guia.id, 'numero': guia.numero, 'nombre': guia.nombre}
        for guia in guias
    ]
    
    return JsonResponse({'guias_laboratorio': guias_data})

@login_required
def get_practicas_ajax(request):
    """Obtener prácticas por guía de laboratorio"""
    guia_id = request.GET.get('guia_id')
    
    if guia_id:
        practicas = Practica.objects.filter(guia_laboratorio_id=guia_id)
    else:
        practicas = Practica.objects.none()
    
    practicas_data = [
        {'id': practica.id, 'numero': practica.numero, 'nombre': practica.nombre}
        for practica in practicas
    ]
    
    return JsonResponse({'practicas': practicas_data})

@login_required
def get_laboratorios_ajax(request):
    """Obtener laboratorios por unidad académica"""
    unidad_id = request.GET.get('unidad_id')
    
    if unidad_id:
        laboratorios = Laboratorio.objects.filter(unidad_academica_id=unidad_id)
    else:
        laboratorios = Laboratorio.objects.all()
    
    laboratorios_data = [
        {
            'id': laboratorio.id, 
            'nombre': laboratorio.nombre, 
            'display': laboratorio.get_nombre_display(),
            'seccion_area': laboratorio.seccion_area,
            'identificador_aula': laboratorio.identificador_aula
        }
        for laboratorio in laboratorios
    ]
    
    return JsonResponse({'laboratorios': laboratorios_data})

@login_required
def filtrar_equipos_ajax(request):
    """Filtrar equipos via AJAX para la vista de lista"""
    
    # Obtener filtros del request
    filtros = {
        'unidad_academica': request.GET.get('unidad_academica', ''),
        'carrera': request.GET.get('carrera', ''),
        'semestre': request.GET.get('semestre', ''),
        'asignatura': request.GET.get('asignatura', ''),
        'estado': request.GET.get('estado', ''),
        'laboratorio': request.GET.get('laboratorio', ''),
        'busqueda': request.GET.get('busqueda', ''),
    }
    
    # Construir queryset
    equipos = Equipo.objects.select_related(
        'unidad_academica', 'carrera', 'asignatura', 'laboratorio',
        'criterio_desempeno', 'unidad_didactica', 'contenido_analitico',
        'guia_laboratorio', 'practica'
    ).all()
    
    if filtros['unidad_academica']:
        equipos = equipos.filter(unidad_academica_id=filtros['unidad_academica'])
    
    if filtros['carrera']:
        equipos = equipos.filter(carrera_id=filtros['carrera'])
    
    if filtros['semestre']:
        equipos = equipos.filter(semestre=filtros['semestre'])
    
    if filtros['asignatura']:
        equipos = equipos.filter(asignatura_id=filtros['asignatura'])
    
    if filtros['estado']:
        equipos = equipos.filter(estado=filtros['estado'])
    
    if filtros['laboratorio']:
        equipos = equipos.filter(laboratorio_id=filtros['laboratorio'])
    
    if filtros['busqueda']:
        equipos = equipos.filter(
            Q(equipo_existente__icontains=filtros['busqueda']) |
            Q(marca__icontains=filtros['busqueda']) |
            Q(modelo__icontains=filtros['busqueda']) |
            Q(codigo_inventario__icontains=filtros['busqueda'])
        )
    
    # Preparar datos para respuesta
    equipos_data = []
    for equipo in equipos[:50]:  # Limitar a 50 resultados
        equipos_data.append({
            'id': equipo.id,
            'codigo_inventario': equipo.codigo_inventario,
            'equipo_existente': equipo.equipo_existente,
            'marca': equipo.marca,
            'modelo': equipo.modelo,
            'estado': equipo.get_estado_display(),
            'unidad_academica': equipo.unidad_academica.get_nombre_display(),
            'carrera': equipo.carrera.get_nombre_display(),
            'semestre': f"{equipo.semestre}° Semestre",
            'asignatura': equipo.asignatura.get_nombre_display(),
            'laboratorio': equipo.laboratorio.get_nombre_display(),
            'numero_unidades': equipo.numero_unidades,
            'created_at': equipo.created_at.strftime('%d/%m/%Y'),
        })
    
    # Estadísticas
    stats = {
        'total': equipos.count(),
        'bueno': equipos.filter(estado='bueno').count(),
        'regular': equipos.filter(estado='regular').count(),
        'malo': equipos.filter(estado='malo').count(),
    }
    
    return JsonResponse({
        'equipos': equipos_data,
        'stats': stats,
        'total_equipos': equipos.count()
    })

@login_required
def exportar_equipos_excel(request):
    """Exportar equipos a Excel con todas las 24 columnas"""
    
    # Obtener filtros si existen
    filtros = {
        'unidad_academica': request.GET.get('unidad_academica', ''),
        'carrera': request.GET.get('carrera', ''),
        'semestre': request.GET.get('semestre', ''),
        'estado': request.GET.get('estado', ''),
    }
    
    # Construir queryset
    equipos = Equipo.objects.select_related(
        'unidad_academica', 'carrera', 'asignatura',
        'guia_laboratorio', 'practica', 'laboratorio', 'usuario_creador'
    ).all()
    
    # Aplicar filtros
    if filtros['unidad_academica']:
        equipos = equipos.filter(unidad_academica_id=filtros['unidad_academica'])
    
    if filtros['carrera']:
        equipos = equipos.filter(carrera_id=filtros['carrera'])
    
    if filtros['semestre']:
        equipos = equipos.filter(semestre=filtros['semestre'])
    
    if filtros['estado']:
        equipos = equipos.filter(estado=filtros['estado'])
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Equipos de Laboratorio"
    
    # Definir estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )
    
    # Encabezados (22 columnas)
    headers = [
        'UNIDAD ACADÉMICA',
        'CARRERA',
        'SEMESTRE',
        'ASIGNATURA',
        'CARGA HORARIA SEMANAL',
        'CARGA HORARIA SEMESTRAL',
        'UNIDAD TEMÁTICA',
        'GUÍA DE LABORATORIO',
        'PRÁCTICA',
        'EQUIPO EXISTENTE',
        'MARCA',
        'MODELO',
        'ESTADO',
        'NÚMERO DE UNIDADES DEL EQUIPO',
        'ES UN ACTIVO FIJO DE ACUERDO A SU ACTA DE ENTREGA?',
        'FOTOGRAFÍA FRONTAL DEL EQUIPO',
        'FOTOGRAFÍA DE LA PLACA DE CARACTERÍSTICAS',
        'UBICACIÓN DEL EQUIPO (LABORATORIO)',
        'SECCIÓN/ÁREA',
        'IDENTIFICADOR/Nº DE AULA',
        'EQUIPO REQUERIDO',
        'NÚMERO DE EQUIPOS REQUERIDOS'
    ]
    
    # Escribir encabezados
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Escribir datos
    for row, equipo in enumerate(equipos, 2):
        data = [
            equipo.unidad_academica.get_nombre_display(),
            equipo.carrera.get_nombre_display(),
            f"{equipo.semestre}° Semestre",
            equipo.asignatura.get_nombre_display(),
            equipo.carga_horaria_semanal,
            equipo.carga_horaria_semestral,
            f"Criterio: {equipo.criterio_desempeno.nombre if equipo.criterio_desempeno else 'N/A'}",
            f"Guía {equipo.guia_laboratorio.numero}: {equipo.guia_laboratorio.nombre}",
            f"Práctica {equipo.practica.numero}: {equipo.practica.nombre}",
            equipo.equipo_existente,
            equipo.marca,
            equipo.modelo,
            equipo.get_estado_display(),
            equipo.numero_unidades,
            'Sí' if equipo.es_activo_fijo else 'No',
            'Sí' if equipo.fotografia_frontal else 'No',
            'Sí' if equipo.fotografia_placa else 'No',
            equipo.laboratorio.get_nombre_display(),
            equipo.seccion_area,
            equipo.identificador_aula,
            equipo.equipo_requerido,
            equipo.numero_equipos_requeridos,
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
    
    # Ajustar ancho de columnas
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 20
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    fecha_actual = timezone.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="equipos_laboratorio_{fecha_actual}.xlsx"'
    
    wb.save(response)
    return response


# ==========================================
# VISTAS PARA SISTEMA DE REORDENAMIENTO
# ==========================================

@login_required
def lista_tareas_reordenamiento(request):
    """Vista principal para mostrar la lista de tareas de reordenamiento"""
    
    # Obtener filtros
    filtros = {
        'estado': request.GET.get('estado', ''),
        'tipo': request.GET.get('tipo', ''),
        'prioridad': request.GET.get('prioridad', ''),
        'usuario_asignado': request.GET.get('usuario_asignado', ''),
        'busqueda': request.GET.get('busqueda', ''),
    }
    
    # Construir queryset con filtros
    tareas = TareaReordenamiento.objects.select_related('usuario_creador', 'usuario_asignado')
    
    if filtros['estado']:
        tareas = tareas.filter(estado=filtros['estado'])
    
    if filtros['tipo']:
        tareas = tareas.filter(tipo=filtros['tipo'])
    
    if filtros['prioridad']:
        tareas = tareas.filter(prioridad=filtros['prioridad'])
    
    if filtros['usuario_asignado']:
        tareas = tareas.filter(usuario_asignado_id=filtros['usuario_asignado'])
    
    if filtros['busqueda']:
        tareas = tareas.filter(
            Q(titulo__icontains=filtros['busqueda']) |
            Q(descripcion__icontains=filtros['busqueda'])
        )
    
    # Paginación
    paginator = Paginator(tareas, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Estadísticas
    stats = {
        'total': TareaReordenamiento.objects.count(),
        'pendientes': TareaReordenamiento.objects.filter(estado='pendiente').count(),
        'en_proceso': TareaReordenamiento.objects.filter(estado='en_proceso').count(),
        'completadas': TareaReordenamiento.objects.filter(estado='completada').count(),
    }
    
    context = {
        'tareas': page_obj,
        'filtros': filtros,
        'stats': stats,
        'tipos_tarea': TareaReordenamiento.TIPOS_TAREA,
        'estados_tarea': TareaReordenamiento.ESTADOS_TAREA,
        'prioridades': TareaReordenamiento.PRIORIDADES,
    }
    
    return render(request, 'equipos/reordenamiento/lista_tareas.html', context)


@login_required
def api_equipos_disponibles(request):
    """API para obtener equipos disponibles para reordenamiento"""
    
    # Obtener parámetros de filtrado
    search = request.GET.get('search', '').strip()
    unidad_id = request.GET.get('unidad_academica', '')
    laboratorio_id = request.GET.get('laboratorio', '')
    estado = request.GET.get('estado', '')
    
    # Construir queryset base
    equipos = Equipo.objects.select_related(
        'unidad_academica', 'carrera', 'laboratorio'
    ).filter(
        estado__in=['operativo', 'mantenimiento', 'reparacion', 'nuevo', 'usado']
    )
    
    # Aplicar filtros
    if search:
        equipos = equipos.filter(
            Q(codigo_inventario__icontains=search) |
            Q(equipo_existente__icontains=search) |
            Q(marca__icontains=search) |
            Q(modelo__icontains=search)
        )
    
    if unidad_id:
        equipos = equipos.filter(unidad_academica_id=unidad_id)
    
    if laboratorio_id:
        equipos = equipos.filter(laboratorio_id=laboratorio_id)
    
    if estado:
        equipos = equipos.filter(estado=estado)
    
    # Limitar resultados
    equipos = equipos[:100]  # Máximo 100 resultados
    
    # Serializar datos
    data = []
    for equipo in equipos:
        data.append({
            'id': equipo.id,
            'codigo_inventario': equipo.codigo_inventario or f'EQ-{equipo.id:04d}',
            'equipo_existente': equipo.equipo_existente,
            'marca': equipo.marca or '',
            'modelo': equipo.modelo or '',
            'estado': equipo.estado,
            'unidad_academica': equipo.unidad_academica.get_nombre_display() if equipo.unidad_academica else '',
            'laboratorio': equipo.laboratorio.get_nombre_display() if equipo.laboratorio else '',
        })
    
    return JsonResponse({
        'equipos': data,
        'total': len(data)
    })


@login_required
def api_laboratorios_por_unidad(request, unidad_id):
    """API para obtener laboratorios (simplificado - todos los laboratorios)"""
    
    try:
        # Por ahora devolvemos todos los laboratorios ya que no hay relación directa
        # En el futuro se podría agregar un campo unidad_academica a Laboratorio
        laboratorios = Laboratorio.objects.all().order_by('nombre')
        
        data = []
        for lab in laboratorios:
            data.append({
                'id': lab.id,
                'nombre': lab.get_nombre_display(),
                'descripcion': lab.descripcion or '',
            })
        
        return JsonResponse({
            'laboratorios': data
        })
        
    except Exception as e:
        return JsonResponse({
            'error': f'Error al cargar laboratorios: {str(e)}'
        }, status=500)


@login_required
def nueva_tarea_reordenamiento(request):
    """Vista para crear una nueva tarea de reordenamiento"""
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Crear la tarea
                tarea = TareaReordenamiento.objects.create(
                    titulo=request.POST.get('titulo'),
                    descripcion=request.POST.get('descripcion'),
                    tipo=request.POST.get('tipo'),
                    prioridad=request.POST.get('prioridad'),
                    fecha_fin_estimada=request.POST.get('fecha_fin_estimada') or None,
                    usuario_creador=request.user,
                    usuario_asignado_id=request.POST.get('usuario_asignado') or None,
                    observaciones=request.POST.get('observaciones', '')
                )
                
                # Procesar equipos seleccionados
                equipos_seleccionados = request.POST.getlist('equipos_seleccionados')
                
                if not equipos_seleccionados:
                    messages.error(request, 'Debe seleccionar al menos un equipo para la tarea.')
                    raise ValueError("No se seleccionaron equipos")
                unidad_destino_id = request.POST.get('unidad_destino')
                laboratorio_destino_id = request.POST.get('laboratorio_destino')
                observaciones_destino = request.POST.get('observaciones_destino', '')
                
                # Obtener objetos de destino si se proporcionaron
                unidad_destino = None
                laboratorio_destino = None
                
                if unidad_destino_id:
                    try:
                        unidad_destino = UnidadAcademica.objects.get(id=unidad_destino_id)
                    except UnidadAcademica.DoesNotExist:
                        pass
                
                if laboratorio_destino_id:
                    try:
                        laboratorio_destino = Laboratorio.objects.get(id=laboratorio_destino_id)
                    except Laboratorio.DoesNotExist:
                        pass
                
                # Crear EquipoTarea para cada equipo seleccionado
                equipos_creados = 0
                for equipo_id in equipos_seleccionados:
                    try:
                        equipo = Equipo.objects.get(id=equipo_id)
                        
                        EquipoTarea.objects.create(
                            tarea=tarea,
                            equipo=equipo,
                            unidad_academica_origen=equipo.unidad_academica,
                            laboratorio_origen=equipo.laboratorio,
                            unidad_academica_destino=unidad_destino,
                            laboratorio_destino=laboratorio_destino,
                            observaciones_equipo=observaciones_destino
                        )
                        equipos_creados += 1
                        
                    except Equipo.DoesNotExist:
                        continue
                
                # Crear log de creación
                LogReordenamiento.objects.create(
                    tarea=tarea,
                    usuario=request.user,
                    accion='Tarea Creada',
                    descripcion=f'Nueva tarea de reordenamiento con {len(equipos_seleccionados)} equipos'
                )
                
                messages.success(request, 'Tarea de reordenamiento creada correctamente.')
                return redirect('equipos:detalle_tarea', pk=tarea.pk)
                
        except Exception as e:
            messages.error(request, f'Error al crear la tarea: {str(e)}')
    
    # Obtener usuarios para asignar
    from django.contrib.auth import get_user_model
    User = get_user_model()
    usuarios = User.objects.filter(is_active=True).order_by('first_name', 'last_name')
    
    # Obtener unidades académicas para el formulario de destino
    unidades_academicas = UnidadAcademica.objects.all().order_by('nombre')
    
    context = {
        'tipos_tarea': TareaReordenamiento.TIPOS_TAREA,
        'prioridades': TareaReordenamiento.PRIORIDADES,
        'usuarios': usuarios,
        'unidades_academicas': unidades_academicas,
    }
    
    return render(request, 'equipos/reordenamiento/nueva_tarea.html', context)


@login_required
def detalle_tarea_reordenamiento(request, pk):
    """Vista detallada de una tarea de reordenamiento"""
    
    tarea = get_object_or_404(TareaReordenamiento, pk=pk)
    
    # Obtener equipos involucrados
    equipos_tarea = EquipoTarea.objects.filter(tarea=tarea).select_related(
        'equipo', 'unidad_academica_origen', 'laboratorio_origen',
        'unidad_academica_destino', 'laboratorio_destino'
    )
    
    # Obtener logs de la tarea
    logs = LogReordenamiento.objects.filter(tarea=tarea).select_related('usuario')
    
    # Estadísticas de la tarea
    total_equipos = equipos_tarea.count()
    equipos_procesados = equipos_tarea.filter(procesado=True).count()
    porcentaje_real = (equipos_procesados / total_equipos * 100) if total_equipos > 0 else 0
    
    context = {
        'tarea': tarea,
        'equipos_tarea': equipos_tarea,
        'logs': logs,
        'total_equipos': total_equipos,
        'equipos_procesados': equipos_procesados,
        'porcentaje_real': porcentaje_real,
    }
    
    return render(request, 'equipos/reordenamiento/detalle_tarea.html', context)


@login_required
def editar_tarea_reordenamiento(request, pk):
    """Vista para editar una tarea de reordenamiento"""
    
    tarea = get_object_or_404(TareaReordenamiento, pk=pk)
    
    if request.method == 'POST':
        try:
            # Guardar estado anterior para el log
            estado_anterior = tarea.estado
            
            # Actualizar la tarea
            tarea.titulo = request.POST.get('titulo')
            tarea.descripcion = request.POST.get('descripcion')
            tarea.tipo = request.POST.get('tipo')
            tarea.estado = request.POST.get('estado')
            tarea.prioridad = request.POST.get('prioridad')
            tarea.fecha_fin_estimada = request.POST.get('fecha_fin_estimada') or None
            tarea.usuario_asignado_id = request.POST.get('usuario_asignado') or None
            tarea.observaciones = request.POST.get('observaciones', '')
            tarea.porcentaje_completado = int(request.POST.get('porcentaje_completado', 0))
            
            # Si se marca como completada, establecer fecha fin real
            if tarea.estado == 'completada' and not tarea.fecha_fin_real:
                tarea.fecha_fin_real = timezone.now()
            
            tarea.save()
            
            # Crear log de modificación
            cambios = []
            if estado_anterior != tarea.estado:
                cambios.append(f'Estado: {estado_anterior} → {tarea.estado}')
            
            if cambios:
                LogReordenamiento.objects.create(
                    tarea=tarea,
                    usuario=request.user,
                    accion='Tarea Modificada',
                    descripcion=f'Cambios: {", ".join(cambios)}'
                )
            
            messages.success(request, 'Tarea actualizada exitosamente.')
            return redirect('equipos:detalle_tarea', pk=tarea.pk)
            
        except Exception as e:
            messages.error(request, f'Error al actualizar la tarea: {str(e)}')
    
    # Obtener usuarios para asignar
    from django.contrib.auth import get_user_model
    User = get_user_model()
    usuarios = User.objects.filter(is_active=True).order_by('first_name', 'last_name')
    
    context = {
        'tarea': tarea,
        'tipos_tarea': TareaReordenamiento.TIPOS_TAREA,
        'estados_tarea': TareaReordenamiento.ESTADOS_TAREA,
        'prioridades': TareaReordenamiento.PRIORIDADES,
        'usuarios': usuarios,
    }
    
    return render(request, 'equipos/reordenamiento/editar_tarea.html', context)


@login_required
def eliminar_tarea_reordenamiento(request, pk):
    """Vista para eliminar una tarea de reordenamiento"""
    
    tarea = get_object_or_404(TareaReordenamiento, pk=pk)
    
    if request.method == 'POST':
        titulo_tarea = tarea.titulo
        tarea.delete()
        messages.success(request, f'Tarea "{titulo_tarea}" eliminada exitosamente.')
        return redirect('equipos:reordenamiento')
    
    context = {
        'tarea': tarea,
    }
    
    return render(request, 'equipos/reordenamiento/eliminar_tarea.html', context)


@login_required
def buscar_equipos_reordenamiento(request):
    """Vista para buscar y seleccionar equipos para reordenamiento"""
    
    tarea_id = request.GET.get('tarea_id')
    tarea = get_object_or_404(TareaReordenamiento, pk=tarea_id) if tarea_id else None
    
    # Obtener filtros de búsqueda
    filtros = {
        'unidad_academica': request.GET.get('unidad_academica', ''),
        'carrera': request.GET.get('carrera', ''),
        'laboratorio': request.GET.get('laboratorio', ''),
        'estado': request.GET.get('estado', ''),
        'busqueda': request.GET.get('busqueda', ''),
    }
    
    # Construir queryset
    equipos = Equipo.objects.select_related(
        'unidad_academica', 'carrera', 'asignatura', 'laboratorio'
    )
    
    if filtros['unidad_academica']:
        equipos = equipos.filter(unidad_academica_id=filtros['unidad_academica'])
    
    if filtros['carrera']:
        equipos = equipos.filter(carrera_id=filtros['carrera'])
    
    if filtros['laboratorio']:
        equipos = equipos.filter(laboratorio_id=filtros['laboratorio'])
    
    if filtros['estado']:
        equipos = equipos.filter(estado=filtros['estado'])
    
    if filtros['busqueda']:
        equipos = equipos.filter(
            Q(equipo_existente__icontains=filtros['busqueda']) |
            Q(marca__icontains=filtros['busqueda']) |
            Q(modelo__icontains=filtros['busqueda']) |
            Q(codigo_inventario__icontains=filtros['busqueda'])
        )
    
    # Si hay una tarea, excluir equipos ya agregados
    if tarea:
        equipos_ya_agregados = EquipoTarea.objects.filter(tarea=tarea).values_list('equipo_id', flat=True)
        equipos = equipos.exclude(id__in=equipos_ya_agregados)
    
    # Paginación
    paginator = Paginator(equipos, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Obtener datos para filtros
    unidades_academicas = UnidadAcademica.objects.all()
    carreras = Carrera.objects.all()
    laboratorios = Laboratorio.objects.all()
    
    context = {
        'equipos': page_obj,
        'tarea': tarea,
        'filtros': filtros,
        'unidades_academicas': unidades_academicas,
        'carreras': carreras,
        'laboratorios': laboratorios,
        'estados': Equipo.ESTADOS,
    }
    
    return render(request, 'equipos/reordenamiento/buscar_equipos.html', context)


@login_required
def procesar_tarea_reordenamiento(request, pk):
    """Vista para procesar/ejecutar una tarea de reordenamiento"""
    
    tarea = get_object_or_404(TareaReordenamiento, pk=pk)
    
    if request.method == 'POST':
        accion = request.POST.get('accion')
        
        if accion == 'agregar_equipos':
            # Agregar equipos seleccionados a la tarea
            equipos_ids = request.POST.getlist('equipos_seleccionados')
            unidad_destino_id = request.POST.get('unidad_destino')
            laboratorio_destino_id = request.POST.get('laboratorio_destino')
            
            try:
                with transaction.atomic():
                    for equipo_id in equipos_ids:
                        equipo = get_object_or_404(Equipo, pk=equipo_id)
                        
                        EquipoTarea.objects.create(
                            tarea=tarea,
                            equipo=equipo,
                            unidad_academica_origen=equipo.unidad_academica,
                            laboratorio_origen=equipo.laboratorio,
                            unidad_academica_destino_id=unidad_destino_id or None,
                            laboratorio_destino_id=laboratorio_destino_id or None,
                        )
                    
                    # Crear log
                    LogReordenamiento.objects.create(
                        tarea=tarea,
                        usuario=request.user,
                        accion='Equipos Agregados',
                        descripcion=f'Se agregaron {len(equipos_ids)} equipos a la tarea'
                    )
                    
                    messages.success(request, f'Se agregaron {len(equipos_ids)} equipos a la tarea.')
                    
            except Exception as e:
                messages.error(request, f'Error al agregar equipos: {str(e)}')
        
        elif accion == 'ejecutar_reordenamiento':
            # Ejecutar el reordenamiento de equipos
            try:
                with transaction.atomic():
                    equipos_tarea = EquipoTarea.objects.filter(tarea=tarea, procesado=False)
                    equipos_procesados = 0
                    
                    for equipo_tarea in equipos_tarea:
                        equipo = equipo_tarea.equipo
                        
                        # Guardar estado anterior para historial
                        estado_anterior = {
                            'unidad_academica': equipo.unidad_academica,
                            'laboratorio': equipo.laboratorio,
                        }
                        
                        # Aplicar cambios según el tipo de tarea
                        if tarea.tipo == 'reasignacion' or tarea.tipo == 'transferencia_unidad':
                            if equipo_tarea.unidad_academica_destino:
                                equipo.unidad_academica = equipo_tarea.unidad_academica_destino
                            if equipo_tarea.laboratorio_destino:
                                equipo.laboratorio = equipo_tarea.laboratorio_destino
                        
                        elif tarea.tipo == 'reubicacion':
                            if equipo_tarea.laboratorio_destino:
                                equipo.laboratorio = equipo_tarea.laboratorio_destino
                        
                        # Guardar equipo
                        equipo.save()
                        
                        # Crear historial
                        HistorialEquipo.objects.create(
                            equipo=equipo,
                            estado_anterior=equipo.estado,
                            estado_nuevo=equipo.estado,
                            usuario=request.user,
                            observaciones=f'Reordenamiento por tarea: {tarea.titulo}'
                        )
                        
                        # Marcar como procesado
                        equipo_tarea.procesado = True
                        equipo_tarea.fecha_procesado = timezone.now()
                        equipo_tarea.save()
                        
                        equipos_procesados += 1
                    
                    # Actualizar estado de la tarea
                    if equipos_procesados > 0:
                        total_equipos = EquipoTarea.objects.filter(tarea=tarea).count()
                        equipos_procesados_total = EquipoTarea.objects.filter(tarea=tarea, procesado=True).count()
                        
                        tarea.porcentaje_completado = int((equipos_procesados_total / total_equipos) * 100)
                        
                        if equipos_procesados_total == total_equipos:
                            tarea.estado = 'completada'
                            tarea.fecha_fin_real = timezone.now()
                        elif tarea.estado == 'pendiente':
                            tarea.estado = 'en_proceso'
                            tarea.fecha_inicio = timezone.now()
                        
                        tarea.save()
                        
                        # Crear log
                        LogReordenamiento.objects.create(
                            tarea=tarea,
                            usuario=request.user,
                            accion='Reordenamiento Ejecutado',
                            descripcion=f'Se procesaron {equipos_procesados} equipos. Progreso: {tarea.porcentaje_completado}%'
                        )
                        
                        messages.success(request, f'Se procesaron {equipos_procesados} equipos exitosamente.')
                    else:
                        messages.warning(request, 'No hay equipos pendientes para procesar.')
                    
            except Exception as e:
                messages.error(request, f'Error al ejecutar reordenamiento: {str(e)}')
        
        return redirect('equipos:detalle_tarea', pk=tarea.pk)
    
    # GET request - mostrar formulario de procesamiento
    equipos_tarea = EquipoTarea.objects.filter(tarea=tarea).select_related('equipo')
    unidades_academicas = UnidadAcademica.objects.all()
    laboratorios = Laboratorio.objects.all()
    
    context = {
        'tarea': tarea,
        'equipos_tarea': equipos_tarea,
        'unidades_academicas': unidades_academicas,
        'laboratorios': laboratorios,
    }
    
    return render(request, 'equipos/reordenamiento/procesar_tarea.html', context)


@csrf_exempt
@login_required
def get_laboratorios_unidad_ajax(request):
    """Obtener laboratorios por unidad académica via AJAX"""
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            unidad_id = data.get('unidad_id')
            
            if unidad_id:
                laboratorios = Laboratorio.objects.filter(unidad_academica_id=unidad_id)
                laboratorios_data = [
                    {
                        'id': lab.id,
                        'nombre': lab.get_nombre_display()
                    }
                    for lab in laboratorios
                ]
                
                return JsonResponse({
                    'success': True,
                    'laboratorios': laboratorios_data
                })
            else:
                return JsonResponse({
                    'success': True,
                    'laboratorios': []
                })
                
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})


# Vistas AJAX para cargar datos dinámicamente en formularios de equipos

@login_required
def cargar_carreras_ajax(request):
    """Cargar carreras basadas en la unidad académica seleccionada"""
    unidad_academica_id = request.GET.get('unidad_academica_id')
    
    if unidad_academica_id:
        carreras = Carrera.objects.filter(unidad_academica_id=unidad_academica_id).values('id', 'nombre')
        carreras_data = [
            {'id': carrera['id'], 'text': dict(Carrera.CARRERAS)[carrera['nombre']]}
            for carrera in carreras
        ]
    else:
        carreras_data = []
    
    return JsonResponse({'carreras': carreras_data})

@login_required
def cargar_asignaturas_ajax(request):
    """Cargar asignaturas basadas en la carrera seleccionada"""
    carrera_id = request.GET.get('carrera_id')
    
    if carrera_id:
        asignaturas = Asignatura.objects.filter(carrera_id=carrera_id).values(
            'id', 'nombre', 'semestre', 'codigo_competencia', 'sigla_curricular',
            'carga_horaria_semanal', 'carga_horaria_semestral'
        )
        asignaturas_data = [
            {
                'id': asignatura['id'], 
                'text': f"{dict(Asignatura.ASIGNATURAS_CHOICES)[asignatura['nombre']]} (Semestre {asignatura['semestre']})",
                'semestre': asignatura['semestre'],
                'codigo_competencia': asignatura['codigo_competencia'] or '',
                'sigla_curricular': asignatura['sigla_curricular'] or '',
                'carga_horaria_semanal': asignatura['carga_horaria_semanal'] or '',
                'carga_horaria_semestral': asignatura['carga_horaria_semestral'] or ''
            }
            for asignatura in asignaturas
        ]
    else:
        asignaturas_data = []
    
    return JsonResponse({'asignaturas': asignaturas_data})

@login_required 
def cargar_criterios_desempeno_ajax(request):
    """Cargar criterios de desempeño basados en la asignatura seleccionada"""
    asignatura_id = request.GET.get('asignatura_id')
    
    if asignatura_id:
        criterios = CriterioDesempeno.objects.filter(asignatura_id=asignatura_id).values(
            'id', 'nombre', 'descripcion'
        )
        criterios_data = [
            {
                'id': criterio['id'],
                'text': criterio['descripcion'],  # Mostrar la descripción completa del Excel
                'descripcion': criterio['descripcion']
            }
            for criterio in criterios
        ]
    else:
        criterios_data = []
    
    return JsonResponse({'criterios': criterios_data})

@login_required
def cargar_unidades_didacticas_ajax(request):
    """Cargar unidades didácticas basadas en la asignatura seleccionada"""
    asignatura_id = request.GET.get('asignatura_id')
    
    if asignatura_id:
        unidades = UnidadDidactica.objects.filter(asignatura_id=asignatura_id).values(
            'id', 'nombre', 'descripcion'
        )
        unidades_data = [
            {
                'id': unidad['id'],
                'text': unidad['descripcion'],  # Mostrar la descripción del Excel
                'descripcion': unidad['descripcion']
            }
            for unidad in unidades
        ]
    else:
        unidades_data = []
    
    return JsonResponse({'unidades_didacticas': unidades_data})

@login_required
def cargar_contenidos_analiticos_ajax(request):
    """Cargar contenidos analíticos basados en la unidad didáctica seleccionada"""
    unidad_didactica_id = request.GET.get('unidad_didactica_id')
    
    if unidad_didactica_id:
        contenidos = ContenidoAnalitico.objects.filter(unidad_didactica_id=unidad_didactica_id).values(
            'id', 'nombre', 'descripcion'
        )
        contenidos_data = [
            {
                'id': contenido['id'],
                'text': contenido['descripcion'],  # Mostrar la descripción del Excel
                'descripcion': contenido['descripcion']
            }
            for contenido in contenidos
        ]
    else:
        contenidos_data = []
    
    return JsonResponse({'contenidos_analiticos': contenidos_data})

@login_required
def cargar_guias_laboratorio_ajax(request):
    """Cargar guías de laboratorio basadas en la asignatura seleccionada"""
    asignatura_id = request.GET.get('asignatura_id')
    
    if asignatura_id:
        guias = GuiaLaboratorio.objects.filter(
            unidad_tematica__asignatura_id=asignatura_id
        ).values('id', 'nombre', 'numero')
        guias_data = [
            {
                'id': guia['id'],
                'text': f"Guía {guia['numero']}: {guia['nombre']}"
            }
            for guia in guias
        ]
    else:
        guias_data = []
    
    return JsonResponse({'guias_laboratorio': guias_data})

@login_required
def cargar_practicas_ajax(request):
    """Cargar prácticas basadas en la guía de laboratorio seleccionada"""
    guia_laboratorio_id = request.GET.get('guia_laboratorio_id')
    
    if guia_laboratorio_id:
        practicas = Practica.objects.filter(guia_laboratorio_id=guia_laboratorio_id).values(
            'id', 'nombre', 'numero'
        )
        practicas_data = [
            {
                'id': practica['id'],
                'text': f"Práctica {practica['numero']}: {practica['nombre']}"
            }
            for practica in practicas
        ]
    else:
        practicas_data = []
    
    return JsonResponse({'practicas': practicas_data})

@login_required
def cargar_equipos_ualp_ajax(request):
    """Cargar equipos disponibles de la UALP desde datos importados"""
    query = request.GET.get('query', '').strip()
    
    # Importar el modelo aquí para evitar problemas de importación circular
    from equipos.models import EquipoImportado
    
    # Filtrar equipos de la UALP
    equipos_query = EquipoImportado.objects.filter(unidad_academica='UALP')
    
    # Si hay un término de búsqueda, filtrar por descripción o código
    if query:
        equipos_query = equipos_query.filter(
            Q(descripcion_activo__icontains=query) | 
            Q(codigo__icontains=query)
        )
    
    # Obtener equipos limitados para evitar sobrecarga
    equipos = equipos_query.values(
        'codigo', 'descripcion_activo', 'responsable', 'estado', 'oficina'
    ).distinct()[:50]
    
    equipos_data = []
    for equipo in equipos:
        # Crear texto descriptivo con la información del Excel
        texto = f"{equipo['codigo']} - {equipo['descripcion_activo']}"
        if equipo['responsable']:
            texto += f" (Resp: {equipo['responsable']})"
        
        equipos_data.append({
            'id': equipo['codigo'],
            'texto': texto,
            'codigo': equipo['codigo'],
            'descripcion': equipo['descripcion_activo'],
            'responsable': equipo['responsable'],
            'estado': equipo['estado'],
            'oficina': equipo['oficina']
        })
    
    return JsonResponse({'equipos': equipos_data})
