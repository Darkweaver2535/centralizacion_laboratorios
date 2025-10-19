from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Count, Sum
from django.core.paginator import Paginator
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
import json
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .models import ActivoFijo
from core.models import UnidadAcademica, Laboratorio, Carrera

@login_required
def seleccionar_metodo_activos(request):
    """Vista de selección de método para gestión de activos fijos"""
    
    # Estadísticas básicas para mostrar en la vista
    stats = {
        'total_activos': ActivoFijo.objects.count(),
        'activos_operativos': ActivoFijo.objects.filter(estado_operativo='operativo').count(),
        'valor_total': ActivoFijo.objects.aggregate(
            total=Sum('valor_adquisicion')
        )['total'] or 0,
        'por_categoria': ActivoFijo.objects.values('categoria').annotate(
            count=Count('id')
        ).order_by('-count')[:5]
    }
    
    context = {
        'stats': stats,
        'page_title': 'Gestión de Activos Fijos',
        'page_description': 'Administración integral de activos fijos institucionales'
    }
    
    return render(request, 'activos/seleccionar_metodo.html', context)

@login_required
def lista_activos(request):
    """Vista principal de activos fijos con filtros y paginación"""
    
    # Obtener parámetros de filtro
    unidad_academica_id = request.GET.get('unidad_academica')
    laboratorio_id = request.GET.get('laboratorio')
    categoria = request.GET.get('categoria')
    estado_fisico = request.GET.get('estado_fisico')
    estado_operativo = request.GET.get('estado_operativo')
    carrera_id = request.GET.get('carrera')
    nombre_activo = request.GET.get('nombre_activo')
    codigo_patrimonial = request.GET.get('codigo_patrimonial')
    
    # Consulta base
    activos = ActivoFijo.objects.select_related(
        'unidad_academica', 'laboratorio', 'carrera', 'creado_por'
    ).all()
    
    # Aplicar filtros
    if unidad_academica_id:
        activos = activos.filter(unidad_academica_id=unidad_academica_id)
    
    if laboratorio_id:
        activos = activos.filter(laboratorio_id=laboratorio_id)
    
    if categoria:
        activos = activos.filter(categoria=categoria)
    
    if estado_fisico:
        activos = activos.filter(estado_fisico=estado_fisico)
        
    if estado_operativo:
        activos = activos.filter(estado_operativo=estado_operativo)
    
    if carrera_id:
        activos = activos.filter(carrera_id=carrera_id)
    
    if nombre_activo:
        activos = activos.filter(
            Q(nombre__icontains=nombre_activo) |
            Q(descripcion__icontains=nombre_activo) |
            Q(marca__icontains=nombre_activo) |
            Q(modelo__icontains=nombre_activo)
        )
        
    if codigo_patrimonial:
        activos = activos.filter(codigo_patrimonial__icontains=codigo_patrimonial)
    
    # Ordenar por código patrimonial
    activos = activos.order_by('codigo_patrimonial')
    
    # Paginación
    paginator = Paginator(activos, 20)  # 20 activos por página
    page_number = request.GET.get('page')
    activos_page = paginator.get_page(page_number)
    
    # Datos para filtros
    unidades_academicas = UnidadAcademica.objects.all()
    laboratorios = Laboratorio.objects.all()
    carreras = Carrera.objects.all()
    
    # Estadísticas para mostrar
    total_activos = activos.count()
    valor_total = activos.aggregate(total=Sum('valor_adquisicion'))['total'] or 0
    
    context = {
        'activos': activos_page,
        'unidades_academicas': unidades_academicas,
        'laboratorios': laboratorios,
        'carreras': carreras,
        'categorias': ActivoFijo.CATEGORIAS,
        'estados_fisicos': ActivoFijo.ESTADOS_FISICOS,
        'estados_operativos': ActivoFijo.ESTADOS_OPERATIVOS,
        
        # Filtros activos
        'unidad_academica_id': unidad_academica_id,
        'laboratorio_id': laboratorio_id,
        'categoria': categoria,
        'estado_fisico': estado_fisico,
        'estado_operativo': estado_operativo,
        'carrera_id': carrera_id,
        'nombre_activo': nombre_activo,
        'codigo_patrimonial': codigo_patrimonial,
        
        # Estadísticas
        'total_activos': total_activos,
        'valor_total': valor_total,
        
        'page_title': 'Lista de Activos Fijos',
    }
    
    return render(request, 'activos/lista_activos.html', context)

@login_required
def detalle_activo(request, activo_id):
    """Vista de detalle de un activo fijo"""
    
    activo = get_object_or_404(ActivoFijo, id=activo_id)
    
    # Calcular depreciación estimada
    valor_depreciado = activo.calcular_depreciacion() if hasattr(activo, 'calcular_depreciacion') else None
    
    context = {
        'activo': activo,
        'valor_depreciado': valor_depreciado,
        'page_title': f'Activo: {activo.codigo_patrimonial}',
    }
    
    return render(request, 'activos/detalle_activo.html', context)

@login_required
def agregar_activo(request):
    """Vista para agregar un nuevo activo fijo"""
    
    if request.method == 'POST':
        try:
            # Crear nuevo activo con los datos del formulario
            activo = ActivoFijo(
                codigo_patrimonial=request.POST.get('codigo_patrimonial'),
                nombre=request.POST.get('nombre'),
                descripcion=request.POST.get('descripcion', ''),
                categoria=request.POST.get('categoria'),
                valor_adquisicion=request.POST.get('valor_adquisicion'),
                fecha_adquisicion=request.POST.get('fecha_adquisicion'),
                metodo_adquisicion=request.POST.get('metodo_adquisicion'),
                proveedor=request.POST.get('proveedor', ''),
                marca=request.POST.get('marca', ''),
                modelo=request.POST.get('modelo', ''),
                numero_serie=request.POST.get('numero_serie', ''),
                año_fabricacion=request.POST.get('año_fabricacion') or None,
                unidad_academica_id=request.POST.get('unidad_academica'),
                laboratorio_id=request.POST.get('laboratorio') or None,
                carrera_id=request.POST.get('carrera') or None,
                ubicacion_fisica=request.POST.get('ubicacion_fisica', ''),
                estado_fisico=request.POST.get('estado_fisico', 'bueno'),
                estado_operativo=request.POST.get('estado_operativo', 'operativo'),
                responsable=request.POST.get('responsable', ''),
                observaciones=request.POST.get('observaciones', ''),
                creado_por=request.user
            )
            
            activo.save()
            
            messages.success(request, f'Activo fijo "{activo.codigo_patrimonial}" creado exitosamente.')
            return redirect('activos:detalle_activo', activo_id=activo.id)
            
        except Exception as e:
            messages.error(request, f'Error al crear el activo fijo: {str(e)}')
    
    # Datos para el formulario
    unidades_academicas = UnidadAcademica.objects.all()
    laboratorios = Laboratorio.objects.all()
    carreras = Carrera.objects.all()
    
    context = {
        'unidades_academicas': unidades_academicas,
        'laboratorios': laboratorios,
        'carreras': carreras,
        'categorias': ActivoFijo.CATEGORIAS,
        'metodos_adquisicion': ActivoFijo.METODOS_ADQUISICION,
        'estados_fisicos': ActivoFijo.ESTADOS_FISICOS,
        'estados_operativos': ActivoFijo.ESTADOS_OPERATIVOS,
        'page_title': 'Agregar Activo Fijo',
    }
    
    return render(request, 'activos/agregar_activo.html', context)

@login_required
def editar_activo(request, activo_id):
    """Vista para editar un activo fijo existente"""
    
    activo = get_object_or_404(ActivoFijo, id=activo_id)
    
    if request.method == 'POST':
        try:
            # Actualizar activo con los datos del formulario
            activo.codigo_patrimonial = request.POST.get('codigo_patrimonial')
            activo.nombre = request.POST.get('nombre')
            activo.descripcion = request.POST.get('descripcion', '')
            activo.categoria = request.POST.get('categoria')
            activo.valor_adquisicion = request.POST.get('valor_adquisicion')
            activo.fecha_adquisicion = request.POST.get('fecha_adquisicion')
            activo.metodo_adquisicion = request.POST.get('metodo_adquisicion')
            activo.proveedor = request.POST.get('proveedor', '')
            activo.marca = request.POST.get('marca', '')
            activo.modelo = request.POST.get('modelo', '')
            activo.numero_serie = request.POST.get('numero_serie', '')
            activo.año_fabricacion = request.POST.get('año_fabricacion') or None
            activo.unidad_academica_id = request.POST.get('unidad_academica')
            activo.laboratorio_id = request.POST.get('laboratorio') or None
            activo.carrera_id = request.POST.get('carrera') or None
            activo.ubicacion_fisica = request.POST.get('ubicacion_fisica', '')
            activo.estado_fisico = request.POST.get('estado_fisico')
            activo.estado_operativo = request.POST.get('estado_operativo')
            activo.responsable = request.POST.get('responsable', '')
            activo.observaciones = request.POST.get('observaciones', '')
            
            activo.save()
            
            messages.success(request, f'Activo fijo "{activo.codigo_patrimonial}" actualizado exitosamente.')
            return redirect('activos:detalle_activo', activo_id=activo.id)
            
        except Exception as e:
            messages.error(request, f'Error al actualizar el activo fijo: {str(e)}')
    
    # Datos para el formulario
    unidades_academicas = UnidadAcademica.objects.all()
    laboratorios = Laboratorio.objects.all()
    carreras = Carrera.objects.all()
    
    context = {
        'activo': activo,
        'unidades_academicas': unidades_academicas,
        'laboratorios': laboratorios,
        'carreras': carreras,
        'categorias': ActivoFijo.CATEGORIAS,
        'metodos_adquisicion': ActivoFijo.METODOS_ADQUISICION,
        'estados_fisicos': ActivoFijo.ESTADOS_FISICOS,
        'estados_operativos': ActivoFijo.ESTADOS_OPERATIVOS,
        'page_title': f'Editar: {activo.codigo_patrimonial}',
    }
    
    return render(request, 'activos/editar_activo.html', context)

@login_required
@require_http_methods(["POST"])
def eliminar_activo(request, activo_id):
    """Vista para eliminar un activo fijo"""
    
    activo = get_object_or_404(ActivoFijo, id=activo_id)
    codigo_patrimonial = activo.codigo_patrimonial
    
    try:
        activo.delete()
        messages.success(request, f'Activo fijo "{codigo_patrimonial}" eliminado exitosamente.')
    except Exception as e:
        messages.error(request, f'Error al eliminar el activo fijo: {str(e)}')
    
    return redirect('activos:lista_activos')

# === VISTAS AJAX ===

@login_required
def ajax_laboratorios_por_unidad(request):
    """API para obtener laboratorios por unidad académica"""
    
    unidad_id = request.GET.get('unidad_id')
    
    if not unidad_id:
        return JsonResponse({'error': 'ID de unidad requerido'}, status=400)
    
    try:
        laboratorios = Laboratorio.objects.filter(
            unidad_academica_id=unidad_id
        ).values('id', 'nombre')
        
        laboratorios_data = [
            {
                'id': lab['id'],
                'nombre': lab['nombre'],
                'display': lab['nombre']
            }
            for lab in laboratorios
        ]
        
        return JsonResponse({
            'laboratorios': laboratorios_data
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def ajax_carreras_por_unidad(request):
    """API para obtener carreras por unidad académica"""
    
    unidad_id = request.GET.get('unidad_id')
    
    if not unidad_id:
        return JsonResponse({'error': 'ID de unidad requerido'}, status=400)
    
    try:
        carreras = Carrera.objects.filter(
            unidad_academica_id=unidad_id
        ).values('id', 'nombre')
        
        carreras_data = [
            {
                'id': carrera['id'],
                'nombre': carrera['nombre'],
                'display': carrera['nombre']
            }
            for carrera in carreras
        ]
        
        return JsonResponse({
            'carreras': carreras_data
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

# === REPORTES Y EXPORTACIÓN ===

@login_required
def exportar_activos_excel(request):
    """Exportar lista de activos fijos a Excel"""
    
    # Aplicar los mismos filtros que en la lista
    activos = ActivoFijo.objects.select_related(
        'unidad_academica', 'laboratorio', 'carrera'
    ).all()
    
    # Aplicar filtros si existen
    unidad_academica_id = request.GET.get('unidad_academica')
    if unidad_academica_id:
        activos = activos.filter(unidad_academica_id=unidad_academica_id)
    
    # Crear workbook de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Activos Fijos"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center")
    
    # Encabezados
    headers = [
        'Código Patrimonial', 'Nombre', 'Categoría', 'Marca', 'Modelo',
        'Valor Adquisición', 'Estado Físico', 'Estado Operativo',
        'Unidad Académica', 'Laboratorio', 'Responsable', 'Fecha Adquisición'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # Datos
    for row, activo in enumerate(activos, 2):
        ws.cell(row=row, column=1, value=activo.codigo_patrimonial)
        ws.cell(row=row, column=2, value=activo.nombre)
        ws.cell(row=row, column=3, value=activo.get_categoria_display())
        ws.cell(row=row, column=4, value=activo.marca)
        ws.cell(row=row, column=5, value=activo.modelo)
        ws.cell(row=row, column=6, value=float(activo.valor_adquisicion))
        ws.cell(row=row, column=7, value=activo.get_estado_fisico_display())
        ws.cell(row=row, column=8, value=activo.get_estado_operativo_display())
        ws.cell(row=row, column=9, value=activo.unidad_academica.get_nombre_display())
        ws.cell(row=row, column=10, value=activo.laboratorio.nombre if activo.laboratorio else '')
        ws.cell(row=row, column=11, value=activo.responsable)
        ws.cell(row=row, column=12, value=activo.fecha_adquisicion.strftime('%d/%m/%Y'))
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="activos_fijos_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    
    wb.save(response)
    return response
